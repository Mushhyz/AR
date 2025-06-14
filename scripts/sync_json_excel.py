"""
Module de synchronisation bidirectionnelle Excel ↔ JSON pour EBIOS RM.
Génère un schéma JSON conforme avec bloc enumerations et metadata.
"""

import json
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class EBIOSJSONExporter:
    """Exporteur JSON conforme EBIOS RM avec métadonnées enrichies."""
    
    def __init__(self, excel_path: Path):
        self.excel_path = Path(excel_path)
        self.wb = None
        
    def load_excel_template(self) -> None:
        """Charge le template Excel EBIOS RM."""
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Template Excel non trouvé : {self.excel_path}")
        
        self.wb = load_workbook(self.excel_path, data_only=True)
        logger.info(f"Template Excel chargé : {self.excel_path}")
    
    def extract_enumerations(self) -> Dict[str, Any]:
        """Extrait toutes les échelles et listes de référence."""
        if not self.wb:
            self.load_excel_template()
        
        # **CORRECTION 1.3** : Bloc enumerations structuré avec libellés français
        enumerations = {
            "gravity_scale": [1, 2, 3, 4],
            "gravity_labels": ["Négligeable", "Limité", "Important", "Critique"],
            "gravity_values": [1, 2, 3, 4],
            "likelihood_scale": [1, 2, 3, 4], 
            "likelihood_labels": ["Minimal", "Significatif", "Élevé", "Maximal"],
            "likelihood_values": [1, 2, 3, 4],
            "business_value_scale": list(range(1, 16)),
            "business_value_labels": [f"Niveau {i}" for i in range(1, 16)],
            "pertinence_scale": [1, 2, 3],
            "pertinence_labels": ["Faible", "Modérée", "Forte"],
            "exposition_scale": [1, 2, 3],
            "exposition_labels": ["Limitée", "Significative", "Maximale"],
            "asset_types": ["Serveur", "Base de données", "Application", "Réseau", "Poste de travail", "Données", "Personnel", "Locaux", "Processus"],
            "stakeholders": ["DSI", "RSSI", "Direction", "DPO", "Métier", "Support", "Externe", "Fournisseur"],
            "treatment_options": ["Réduire", "Éviter", "Transférer", "Accepter"],
            "measure_categories": ["Organisationnelles", "Personnel", "Physiques", "Techniques", "Juridiques"],
            "risk_levels": ["Faible", "Moyen", "Élevé", "Critique"],
            "implementation_status": ["Planifiée", "En cours", "Terminée", "Reportée", "Annulée"]
        }
        
        return enumerations
    
    def extract_measure_catalog(self) -> List[Dict[str, Any]]:
        """Extrait le catalogue des mesures ISO 27001."""
        if not self.wb:
            self.load_excel_template()
        
        # **CORRECTION 6** : Catalogue des mesures avec mapping Annex A
        measures = []
        refs_ws = self.wb["__REFS"]
        
        # Rechercher la table tbl_Measure dans __REFS
        measure_start_row = self._find_table_start(refs_ws, "tbl_Measure")
        if measure_start_row:
            row = measure_start_row + 1  # Ignorer l'en-tête
            while refs_ws.cell(row=row, column=1).value:  # Mesure_ID colonne A
                measure = {
                    "id": refs_ws.cell(row=row, column=1).value,
                    "label": refs_ws.cell(row=row, column=2).value,
                    "category": refs_ws.cell(row=row, column=3).value,
                    "implementation_cost": refs_ws.cell(row=row, column=4).value,
                    "effectiveness_pct": refs_ws.cell(row=row, column=5).value,
                    "annex_a_control": refs_ws.cell(row=row, column=6).value,
                    "iso27001_domain": self._get_iso_domain(refs_ws.cell(row=row, column=6).value)
                }
                measures.append(measure)
                row += 1
        
        return measures
    
    def extract_ebios_data(self) -> Dict[str, Any]:
        """Extrait toutes les données EBIOS RM dans un format JSON structuré."""
        if not self.wb:
            self.load_excel_template()
        
        # **CORRECTION 5** : Structure JSON complète avec bloc enumerations
        ebios_data = {
            "metadata": {
                "generator": "EBIOS RM Template Generator",
                "version": "2.1.0",
                "generated_at": datetime.now().isoformat(),
                "methodology": "EBIOS Risk Manager (ANSSI)",
                "compliance": ["ISO 27001:2022", "ISO 27005:2022", "ISO 31000:2018"],
                "total_assets": self._count_non_empty_rows("Atelier1_Socle"),
                "total_scenarios": self._count_non_empty_rows("Atelier3_Scenarios"),
                "total_measures": self._count_non_empty_rows("Atelier5_Traitement")
            },
            
            # **CORRECTION 5** : Bloc enumerations en racine
            "enumerations": self.extract_enumerations(),
            
            "measure_catalog": self.extract_measure_catalog(),
            
            # **CORRECTION 5** : SoA généré automatiquement
            "annexa_controls": self._generate_annexa_soa(),
            
            "assets": self._extract_sheet_data("Atelier1_Socle", [
                "ID_Actif", "Type", "Sous_Type", "Libellé", "Description",
                "Gravité", "Confidentialité", "Intégrité", "Disponibilité",
                "Valeur_Métier", "Propriétaire", "Score_Risque"
            ]),
            
            "risk_sources": self._extract_sheet_data("Atelier2_Sources", [
                "ID_Source", "Libellé", "Catégorie", "Motivation_Ressources",
                "Ciblage", "Pertinence", "Exposition", "Commentaires"
            ]),
            
            "strategic_scenarios": self._extract_sheet_data("Atelier3_Scenarios", [
                "ID_Scénario", "Source_Risque", "Objectif_Visé", "Chemin_Attaque",
                "Motivation", "Gravité", "Vraisemblance", "Valeur_Métier", "Risque_Calculé"
            ]),
            
            "operational_scenarios": self._extract_sheet_data("Atelier4_Operationnels", [
                "ID_OV", "Scénario_Stratégique", "Vecteur_Attaque", "Étapes_Opérationnelles",
                "Contrôles_Existants", "Vraisemblance_Résiduelle", "Impact", "Risque_Initial",
                "Mesures_Appliquées", "Efficacité_Totale", "Risque_Résiduel", "Niveau_Risque_Final"
            ]),
            
            "treatment_plan": self._extract_sheet_data("Atelier5_Traitement", [
                "ID_Risque", "Scénario_Lié", "Niveau_Initial", "Option_Traitement",
                "Mesure_Choisie", "Contrôle_AnnexA", "Responsable", "Échéance",
                "Coût_Estimé", "Efficacité_Attendue", "Niveau_Résiduel", "Statut_Mise_en_Œuvre"
            ])
        }
        
        return ebios_data
    
    def _extract_column_values(self, ws, column: str, context: str) -> List[str]:
        """Extrait les valeurs d'une colonne en ignorant les en-têtes."""
        values = []
        if column:
            row = 2  # Ignorer l'en-tête
            while ws[f"{column}{row}"].value:
                values.append(str(ws[f"{column}{row}"].value))
                row += 1
        return values
    
    def _find_table_start(self, ws, table_name: str) -> Optional[int]:
        """Trouve la ligne de début d'une table dans __REFS."""
        for row in range(1, 100):  # Recherche dans les 100 premières lignes
            for col in range(1, 50):  # Recherche dans les 50 premières colonnes
                if ws.cell(row=row, column=col).value == table_name:
                    return row
        return None
    
    def _count_non_empty_rows(self, sheet_name: str) -> int:
        """Compte les lignes non vides d'une feuille (hors en-tête)."""
        if sheet_name not in self.wb.sheetnames:
            return 0
        
        ws = self.wb[sheet_name]
        count = 0
        row = 2  # Ignorer l'en-tête
        while ws.cell(row=row, column=1).value:  # Première colonne = ID
            count += 1
            row += 1
        return count
    
    def _extract_sheet_data(self, sheet_name: str, headers: List[str]) -> List[Dict[str, Any]]:
        """Extrait les données d'une feuille sous forme de liste de dictionnaires."""
        if sheet_name not in self.wb.sheetnames:
            return []
        
        ws = self.wb[sheet_name]
        data = []
        
        row = 2  # Ignorer l'en-tête
        while ws.cell(row=row, column=1).value:  # Première colonne = ID
            item = {}
            for col, header in enumerate(headers, 1):
                value = ws.cell(row=row, column=col).value
                item[header.lower().replace("_", "")] = value if value is not None else ""
            data.append(item)
            row += 1
        
        return data
    
    def _generate_annexa_soa(self) -> Dict[str, Any]:
        """Génère la Statement of Applicability (SoA) ISO 27001."""
        # **CORRECTION 5** : SoA automatique basée sur les mesures sélectionnées
        measures = self.extract_measure_catalog()
        
        # Grouper par domaine ISO 27001
        domains = {}
        for measure in measures:
            domain = measure.get("iso27001_domain", "Autres")
            if domain not in domains:
                domains[domain] = {
                    "controls_count": 0,
                    "implemented_count": 0,
                    "controls": []
                }
            
            domains[domain]["controls_count"] += 1
            domains[domain]["controls"].append({
                "id": measure["id"],
                "label": measure["label"],
                "implementation_status": "Implémenté" if measure.get("effectiveness_pct", 0) > 0 else "Non applicable",
                "justification": f"Efficacité évaluée à {measure.get('effectiveness_pct', 0)}%",
                "implementation_notes": "À compléter par l'organisation",
                "residual_risk": "Acceptable" if measure.get("effectiveness_pct", 0) >= 80 else "À traiter"
            })
            
            if measure.get("effectiveness_pct", 0) > 0:
                domains[domain]["implemented_count"] += 1
        
        # Calculer la couverture globale
        total_controls = sum(d["controls_count"] for d in domains.values())
        total_implemented = sum(d["implemented_count"] for d in domains.values())
        
        soa = {
            "iso27001_version": "2022",
            "assessment_date": datetime.now().strftime("%Y-%m-%d"),
            "overall_coverage": round((total_implemented / total_controls) * 100, 1) if total_controls > 0 else 0,
            "domains": domains,
            "summary": {
                "total_controls": total_controls,
                "implemented": total_implemented,
                "not_applicable": total_controls - total_implemented,
                "coverage_target": 90,
                "compliance_level": "Partiel" if total_implemented / total_controls < 0.9 else "Conforme"
            },
            "export_notes": "SoA générée automatiquement depuis le template EBIOS RM",
            "next_review_date": "À définir par l'organisation"
        }
        
        return soa
    
    def export_soa_file(self, output_path: Path) -> None:
        """Exporte le Statement of Applicability ISO 27001 en fichier dédié."""
        soa_data = self._generate_annexa_soa()
        
        # **CORRECTION 5** : Fichier SoA.xlsx séparé
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "SoA_ISO27001"
        
        # Titre
        ws["A1"] = "STATEMENT OF APPLICABILITY ISO 27001:2022"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:G1")
        
        # En-têtes
        headers = ["Contrôle", "Libellé", "Domaine", "Statut", "Justification", "Risque Résiduel", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Données
        row = 4
        for domain_name, domain_data in soa_data["domains"].items():
            for control in domain_data["controls"]:
                ws.cell(row=row, column=1, value=control["id"])
                ws.cell(row=row, column=2, value=control["label"])
                ws.cell(row=row, column=3, value=domain_name)
                ws.cell(row=row, column=4, value=control["implementation_status"])
                ws.cell(row=row, column=5, value=control["justification"])
                ws.cell(row=row, column=6, value=control["residual_risk"])
                ws.cell(row=row, column=7, value=control["implementation_notes"])
                row += 1
        
        # Métadonnées
        ws["A1"] = f"Couverture globale : {soa_data['overall_coverage']}%"
        ws["A2"] = f"Date d'évaluation : {soa_data['assessment_date']}"
