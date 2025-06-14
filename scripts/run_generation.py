"""
Script orchestrateur pour la génération complète du template EBIOS RM.
Gère les erreurs et fournit un feedback détaillé.
"""

import sys
from pathlib import Path
import logging
from datetime import datetime

def setup_logging():
    """Configure le logging avec sortie console et fichier."""
    log_dir = Path("c:/Users/mushm/Documents/AR/logs")
    log_dir.mkdir(parents=True, exist_ok=True)
    
    log_file = log_dir / f"generation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    return log_file

def check_prerequisites():
    """Vérifie les prérequis avant génération."""
    print("🔍 Vérification des prérequis...")
    
    try:
        import openpyxl
        print(f"✅ OpenPyXL version {openpyxl.__version__}")
    except ImportError:
        print("❌ OpenPyXL non installé")
        print("💡 Exécutez: pip install openpyxl")
        return False
    
    # Vérifier les permissions d'écriture
    test_dir = Path("c:/Users/mushm/Documents/AR/templates")
    try:
        test_dir.mkdir(parents=True, exist_ok=True)
        test_file = test_dir / "test_permissions.tmp"
        test_file.write_text("test")
        test_file.unlink()
        print("✅ Permissions d'écriture OK")
    except Exception as e:
        print(f"❌ Problème de permissions : {e}")
        return False
    
    return True

def run_template_generation():
    """Exécute la génération du template avec gestion d'erreurs."""
    print("\n🚀 GÉNÉRATION DU TEMPLATE EBIOS RM")
    print("=" * 60)
    
    try:
        # Import du générateur
        print("📦 Import du générateur...")
        from generate_template import EBIOSTemplateGenerator
        
        # Initialisation
        print("🔧 Initialisation...")
        generator = EBIOSTemplateGenerator()
        
        # Définition du chemin de sortie
        output_path = Path("c:/Users/mushm/Documents/AR/templates/ebios_risk_assessment_FR.xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        print(f"📁 Destination : {output_path}")
        
        # Génération
        print("⚙️  Génération en cours...")
        print("   • Création des tables de référence...")
        print("   • Configuration des onglets EBIOS RM...")
        print("   • Application des validations...")
        print("   • Formatage et protection...")
        
        generator.generate_template(output_path)
        
        # Vérification du résultat
        if output_path.exists():
            file_size = output_path.stat().st_size / 1024
            print(f"\n✅ GÉNÉRATION RÉUSSIE!")
            print(f"📁 Fichier : {output_path}")
            print(f"📊 Taille : {file_size:.1f} KB")
            
            # Validation du contenu
            try:
                from openpyxl import load_workbook
                wb = load_workbook(output_path)
                sheets = wb.sheetnames
                wb.close()
                
                print(f"📋 Onglets créés ({len(sheets)}) :")
                for sheet in sheets:
                    print(f"   • {sheet}")
                
            except Exception as e:
                print(f"⚠️  Impossible de valider le contenu : {e}")
            
            return output_path
        else:
            print("❌ Le fichier n'a pas été créé")
            return None
            
    except ImportError as e:
        print(f"❌ Erreur d'import : {e}")
        print("💡 Vérifiez que generate_template.py est présent")
        return None
    except Exception as e:
        print(f"❌ Erreur de génération : {e}")
        logging.exception("Erreur détaillée")
        return None

def run_visualization(template_path):
    """Exécute la génération des visualisations."""
    print("\n🎨 GÉNÉRATION DES VISUALISATIONS")
    print("=" * 60)
    
    try:
        from visualize_template import EBIOSVisualizationEngine
        
        print("🔧 Initialisation du moteur de visualisation...")
        visualizer = EBIOSVisualizationEngine(template_path)
        
        print("📊 Chargement du template...")
        visualizer.load_template()
        
        print("🎨 Génération des visualisations...")
        success = visualizer.generate_all_visualizations()
        
        if success:
            print("✅ VISUALISATIONS CRÉÉES!")
            print("📊 Nouveaux onglets :")
            print("   • HeatMap_Risques")
            print("   • TCD_Risques_Proprietaire") 
            print("   • Analyse_AnnexA")
            print("   • Tendances_Evolutives")
            print("   • Resume_Executif")
            return True
        else:
            print("❌ Échec de la génération des visualisations")
            return False
            
    except ImportError as e:
        print(f"❌ Erreur d'import visualisation : {e}")
        return False
    except Exception as e:
        print(f"❌ Erreur de visualisation : {e}")
        logging.exception("Erreur détaillée")
        return False

def main():
    """Point d'entrée principal."""
    print("🎯 GÉNÉRATEUR COMPLET EBIOS RM")
    print("=" * 60)
    print(f"🕐 Début : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Configuration du logging
    log_file = setup_logging()
    print(f"📝 Log détaillé : {log_file}")
    
    # Vérification des prérequis
    if not check_prerequisites():
        print("\n❌ Prérequis non satisfaits - arrêt")
        return False
    
    # Génération du template
    template_path = run_template_generation()
    if not template_path:
        print("\n❌ Génération du template échouée")
        return False
    
    # Génération des visualisations
    viz_success = run_visualization(template_path)
    
    # Résumé final
    print("\n" + "=" * 60)
    print("🎯 RÉSUMÉ FINAL")
    print("=" * 60)
    print(f"✅ Template principal : {template_path}")
    print(f"{'✅' if viz_success else '⚠️ '} Visualisations : {'OK' if viz_success else 'Partielles'}")
    print(f"🕐 Fin : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    if viz_success:
        print("\n🎉 GÉNÉRATION COMPLÈTE RÉUSSIE!")
        print("💡 Le template EBIOS RM est maintenant prêt à l'emploi")
    else:
        print("\n⚠️  GÉNÉRATION PARTIELLE")
        print("💡 Le template de base est créé, mais les visualisations ont échoué")
    
    return True

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
