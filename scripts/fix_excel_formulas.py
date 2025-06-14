"""
Script de diagnostic et correction des formules Excel corrompues.
Répare les erreurs de formule dans les templates EBIOS RM.
"""

from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def diagnose_excel_file(file_path: Path) -> bool:
    """Diagnostique un fichier Excel pour détecter les erreurs."""
    
    print(f"🔍 DIAGNOSTIC du fichier : {file_path}")
    print("="*50)
    
    try:
        # Tentative de chargement
        wb = load_workbook(file_path, data_only=False)
        print("✅ Fichier chargé avec succès")
        
        # Vérification des feuilles
        sheets = wb.sheetnames
        print(f"📄 Nombre de feuilles : {len(sheets)}")
        
        for sheet_name in sheets:
            print(f"   • {sheet_name}")
            
            ws = wb[sheet_name]
            
            # Recherche de formules potentiellement problématiques
            problematic_formulas = []
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if cell.value.startswith('='):
                            # Vérifier si la formule contient des références problématiques
                            formula = cell.value
                            if any(issue in formula for issue in ['PI()', 'RAND()', 'SIN(']):
                                problematic_formulas.append((cell.coordinate, formula))
            
            if problematic_formulas:
                print(f"   ⚠️  {len(problematic_formulas)} formule(s) problématique(s) détectée(s)")
                for coord, formula in problematic_formulas[:3]:  # Afficher les 3 premières
                    print(f"      {coord}: {formula[:50]}...")
        
        return True
        
    except Exception as e:
        print(f"❌ Erreur lors du diagnostic : {e}")
        return False

def repair_excel_file(file_path: Path) -> bool:
    """Répare un fichier Excel en supprimant les formules problématiques."""
    
    print(f"🔧 RÉPARATION du fichier : {file_path}")
    
    try:
        # Créer une sauvegarde
        backup_path = file_path.with_suffix('.xlsx.corrupted')
        if file_path.exists():
            file_path.rename(backup_path)
            print(f"💾 Sauvegarde créée : {backup_path}")
        
        # Créer un nouveau fichier propre
        create_clean_template(file_path)
        
        print("✅ Fichier réparé avec succès")
        return True
        
    except Exception as e:
        print(f"❌ Erreur lors de la réparation : {e}")
        return False

def create_clean_template(output_path: Path) -> None:
    """Crée un template EBIOS RM propre sans formules problématiques."""
    
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Créer les onglets principaux
    create_clean_atelier1(wb)
    create_clean_atelier2(wb) 
    create_clean_atelier3(wb)
    create_clean_atelier4(wb)
    create_clean_config(wb)
    
    # Sauvegarder
    wb.save(output_path)
    print(f"✅ Template propre créé : {output_path}")

def create_clean_atelier1(wb) -> None:
    """Crée l'Atelier 1 sans formules problématiques."""
    ws = wb.create_sheet("Atelier1_Socle")
    
    # En-têtes
    headers = ["ID_Actif", "Type", "Libellé", "Description", "Gravité",
               "Confidentialité", "Intégrité", "Disponibilité", 
               "Valeur_Métier", "Propriétaire", "Score_Risque"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Données d'exemple avec scores calculés statiquement
    sample_data = [
        ["A001", "Serveur", "Serveur web principal", "Serveur hébergeant l'application web", "Important", "Important", "Important", "Critique", "10", "DSI", 64],
        ["A002", "Base de données", "Base clients", "Base de données des clients", "Critique", "Critique", "Important", "Important", "12", "RSSI", 96],
        ["A003", "Application", "ERP", "Système de gestion intégré", "Important", "Limité", "Important", "Important", "8", "Métier", 32],
        ["A004", "Réseau", "Infrastructure réseau", "Équipements réseau principaux", "Important", "Important", "Critique", "Critique", "9", "DSI", 72],
        ["A005", "Poste de travail", "Postes utilisateurs", "Ordinateurs des employés", "Limité", "Limité", "Limité", "Important", "6", "Support", 18],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

def create_clean_atelier2(wb) -> None:
    """Crée l'Atelier 2 sans formules problématiques."""
    ws = wb.create_sheet("Atelier2_Sources")
    
    headers = ["ID_Source", "Libellé", "Catégorie", "Motivation_Ressources", 
               "Ciblage", "Pertinence", "Exposition", "Commentaires"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    sample_data = [
        ["RS001", "Cybercriminels organisés", "Criminalité organisée", "Gain financier", "Données sensibles", "Forte", "Significative", "Menace principale"],
        ["RS002", "Employés malveillants", "Menace interne", "Vengeance", "Systèmes internes", "Modérée", "Limitée", "Risque modéré"],
        ["RS003", "Acteurs étatiques", "Espionnage", "Intelligence", "Propriété intellectuelle", "Forte", "Significative", "APT ciblée"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

def create_clean_atelier3(wb) -> None:
    """Crée l'Atelier 3 sans formules problématiques."""
    ws = wb.create_sheet("Atelier3_Scenarios")
    
    headers = ["ID_Scénario", "Source_Risque", "Objectif_Visé", "Chemin_Attaque",
               "Motivation", "Gravité", "Vraisemblance", "Valeur_Métier", "Risque_Calculé"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    sample_data = [
        ["SR001", "RS001", "Vol de données clients", "Attaque externe ciblée", "Revente de données", 3, 3, 10, 90],
        ["SR002", "RS002", "Sabotage système", "Abus de privilèges", "Vengeance", 4, 2, 8, 64],
        ["SR003", "RS003", "Espionnage industriel", "APT avancée", "Avantage concurrentiel", 3, 3, 12, 108],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

def create_clean_atelier4(wb) -> None:
    """Crée l'Atelier 4 sans formules problématiques."""
    ws = wb.create_sheet("Atelier4_Operationnels")
    
    headers = ["ID_OV", "Scénario_Stratégique", "Vecteur_Attaque", "Étapes_Opérationnelles",
               "Contrôles_Existants", "Vraisemblance_Résiduelle", "Impact", "Niveau_Risque"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    sample_data = [
        ["OV001", "SR001", "Phishing ciblé", "Reconnaissance > Intrusion > Exfiltration", "Formation, antivirus", 3, 3, "Critique"],
        ["OV002", "SR002", "Accès physique", "Planification > Exécution > Destruction", "Contrôle d'accès physique", 2, 4, "Élevé"],
        ["OV003", "SR003", "Infiltration APT", "Infection > Persistance > Collecte > Exfiltration", "EDR, monitoring", 3, 3, "Critique"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

def create_clean_config(wb) -> None:
    """Crée l'onglet de configuration."""
    ws = wb.create_sheet("Config_EBIOS", 0)  # Première position
    
    ws["A1"] = "🔧 CONFIGURATION EBIOS RISK MANAGER"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws.merge_cells("A1:F1")
    
    ws["A3"] = "📋 INSTRUCTIONS D'UTILISATION"
    ws["A3"].font = Font(size=12, bold=True)
    
    instructions = [
        "1. Renseignez les actifs dans l'Atelier 1 - Socle",
        "2. Analysez les sources de risque dans l'Atelier 2", 
        "3. Définissez les scénarios dans l'Atelier 3",
        "4. Évaluez les mesures dans l'Atelier 4",
        "5. Consultez les visualisations pour l'analyse"
    ]
    
    for i, instruction in enumerate(instructions, 5):
        ws.cell(row=i, column=1, value=instruction)

def main():
    """Point d'entrée principal."""
    template_path = Path("c:/Users/mushm/Documents/AR/templates/ebios_risk_assessment_FR.xlsx")
    
    print("🩺 DIAGNOSTIC ET RÉPARATION EXCEL")
    print("="*50)
    
    if template_path.exists():
        # Diagnostic
        is_valid = diagnose_excel_file(template_path)
        
        if not is_valid:
            print("\n🔧 Lancement de la réparation...")
            success = repair_excel_file(template_path)
            
            if success:
                print("\n✅ Réparation terminée avec succès!")
                print(f"📁 Nouveau fichier propre : {template_path}")
            else:
                print("\n❌ Échec de la réparation")
        else:
            print("\n✅ Le fichier semble correct")
    else:
        print(f"📁 Fichier non trouvé, création d'un nouveau template...")
        create_clean_template(template_path)

if __name__ == "__main__":
    main()
