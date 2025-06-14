"""
Générateur de template Excel EBIOS RM conforme aux spécifications.
Crée l'onglet __REFS avec toutes les tables de référence et plages nommées.
Applique les validations de données selon la méthodologie EBIOS RM.
"""

from pathlib import Path
from typing import Dict, List, Any
import logging

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class EBIOSTemplateGenerator:
    """Générateur de template Excel EBIOS RM avec validation complète."""
    
    def __init__(self):
        """Initialise le générateur avec les styles et données de référence."""
        self.wb = Workbook()
        # Styles pour le formatage
        self.gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        
        # Données de référence EBIOS RM conformes
        self.reference_data = self._get_ebios_reference_data()
    
    def _get_ebios_reference_data(self) -> Dict[str, List[Dict[str, Any]]]:
        """Définit les données de référence EBIOS RM conformes à la méthodologie."""
        return {
            # Table des niveaux de gravité (échelle 4 niveaux)
            "tbl_Gravite": [
                {"ID": 1, "Libelle": "Négligeable", "Valeur": 1},
                {"ID": 2, "Libelle": "Limité", "Valeur": 2},
                {"ID": 3, "Libelle": "Important", "Valeur": 3},
                {"ID": 4, "Libelle": "Critique", "Valeur": 4}
            ],
            
            # Table des niveaux de vraisemblance (échelle 4 niveaux)
            "tbl_Vraisemblance": [
                {"ID": 1, "Libelle": "Minimal", "Valeur": 1},
                {"ID": 2, "Libelle": "Significatif", "Valeur": 2},
                {"ID": 3, "Libelle": "Élevé", "Valeur": 3},
                {"ID": 4, "Libelle": "Maximal", "Valeur": 4}
            ],
            
            # **CORRECTION 1.2** : Tables dédiées pour Pertinence/Exposition (échelle 1-3)
            "tbl_Pertinence": [
                {"ID": 1, "Libelle": "Faible", "Valeur": 1},
                {"ID": 2, "Libelle": "Modérée", "Valeur": 2},
                {"ID": 3, "Libelle": "Forte", "Valeur": 3}
            ],
            
            "tbl_Exposition": [
                {"ID": 1, "Libelle": "Limitée", "Valeur": 1},
                {"ID": 2, "Libelle": "Significative", "Valeur": 2},
                {"ID": 3, "Libelle": "Maximale", "Valeur": 3}
            ],
            
            # **CORRECTION 1.2** : Table des valeurs métier avec dissociation ID/Valeur
            "tbl_ValeurMetier": [
                {"ID": i, "Libelle": f"Niveau {i}", "Description": f"Valeur métier niveau {i}", "Valeur": i}
                for i in range(1, 16)
            ],

            # **CORRECTION 1.1** : Catalogue complet des mesures ISO 27001:2022 Annex A
            "tbl_Measure": [
                {"Measure_ID": "A.5.1", "Libelle": "Politiques de sécurité de l'information", "Category": "Organisationnelles", "Cout": 2, "Efficacite_pct": 80, "AnnexA_Control": "A.5.1"},
                {"Measure_ID": "A.5.2", "Libelle": "Rôles et responsabilités en matière de sécurité", "Category": "Organisationnelles", "Cout": 1, "Efficacite_pct": 70, "AnnexA_Control": "A.5.2"},
                {"Measure_ID": "A.5.3", "Libelle": "Séparation des tâches", "Category": "Organisationnelles", "Cout": 2, "Efficacite_pct": 85, "AnnexA_Control": "A.5.3"},
                {"Measure_ID": "A.6.1", "Libelle": "Criblage des antécédents", "Category": "Personnel", "Cout": 2, "Efficacite_pct": 75, "AnnexA_Control": "A.6.1"},
                {"Measure_ID": "A.6.3", "Libelle": "Sensibilisation et formation à la sécurité", "Category": "Personnel", "Cout": 3, "Efficacite_pct": 90, "AnnexA_Control": "A.6.3"},
                {"Measure_ID": "A.7.1", "Libelle": "Sécurité physique des zones", "Category": "Physiques", "Cout": 4, "Efficacite_pct": 95, "AnnexA_Control": "A.7.1"},
                {"Measure_ID": "A.8.1", "Libelle": "Inventaire des actifs", "Category": "Techniques", "Cout": 2, "Efficacite_pct": 85, "AnnexA_Control": "A.8.1"},
                {"Measure_ID": "A.8.5", "Libelle": "Classification de l'information", "Category": "Techniques", "Cout": 2, "Efficacite_pct": 85, "AnnexA_Control": "A.8.5"},
                {"Measure_ID": "A.9.1", "Libelle": "Politique de contrôle d'accès", "Category": "Techniques", "Cout": 2, "Efficacite_pct": 85, "AnnexA_Control": "A.9.1"},
                {"Measure_ID": "A.9.3", "Libelle": "Gestion des comptes d'utilisateur privilégiés", "Category": "Techniques", "Cout": 4, "Efficacite_pct": 95, "AnnexA_Control": "A.9.3"},
                {"Measure_ID": "A.10.1", "Libelle": "Politique d'utilisation des contrôles cryptographiques", "Category": "Techniques", "Cout": 2, "Efficacite_pct": 85, "AnnexA_Control": "A.10.1"},
                {"Measure_ID": "A.11.1", "Libelle": "Procédures d'exploitation sécurisées", "Category": "Techniques", "Cout": 3, "Efficacite_pct": 80, "AnnexA_Control": "A.11.1"},
                {"Measure_ID": "A.12.1", "Libelle": "Procédures d'exploitation sécurisées", "Category": "Techniques", "Cout": 3, "Efficacite_pct": 80, "AnnexA_Control": "A.12.1"},
                {"Measure_ID": "A.13.1", "Libelle": "Contrôles de sécurité dans l'analyse et la spécification", "Category": "Techniques", "Cout": 3, "Efficacite_pct": 75, "AnnexA_Control": "A.13.1"},
                {"Measure_ID": "A.14.1", "Libelle": "Gestion des événements de sécurité de l'information", "Category": "Organisationnelles", "Cout": 3, "Efficacite_pct": 85, "AnnexA_Control": "A.14.1"},
                {"Measure_ID": "A.15.1", "Libelle": "Gestion de la continuité de la sécurité de l'information", "Category": "Organisationnelles", "Cout": 4, "Efficacite_pct": 90, "AnnexA_Control": "A.15.1"},
                {"Measure_ID": "A.16.1", "Libelle": "Conformité aux exigences légales et contractuelles", "Category": "Juridiques", "Cout": 2, "Efficacite_pct": 70, "AnnexA_Control": "A.16.1"}
            ],
            
            # **CORRECTION 3.2** : Table des KPI Velocity/Preparedness ISO 27005:2022
            "tbl_KPI": [
                {"KPI_ID": "VEL001", "Libelle": "Velocity Detection", "Category": "Velocity", "Target": 24, "Unit": "heures", "Scale": 4},
                {"KPI_ID": "VEL002", "Libelle": "Velocity Response", "Category": "Velocity", "Target": 4, "Unit": "heures", "Scale": 4},
                {"KPI_ID": "PREP001", "Libelle": "Preparedness Coverage", "Category": "Preparedness", "Target": 95, "Unit": "%", "Scale": 4},
                {"KPI_ID": "PREP002", "Libelle": "Preparedness Training", "Category": "Preparedness", "Target": 90, "Unit": "%", "Scale": 4}
            ],
            
            # Catalogue des sources de risque EBIOS RM
            "tbl_Source": [
                {
                    "Source_ID": "RS001",
                    "Label": "Cybercriminels organisés",
                    "Category": "Criminalité organisée",
                    "MotivationResources": "Gain financier - Outils avancés",
                    "Targeting": "Données sensibles et systèmes de paiement"
                },
                {
                    "Source_ID": "RS002", 
                    "Label": "Acteurs étatiques",
                    "Category": "Espionnage d'État",
                    "MotivationResources": "Intelligence économique - Ressources illimitées",
                    "Targeting": "Informations stratégiques et propriété intellectuelle"
                },
                {
                    "Source_ID": "RS003",
                    "Label": "Employés malveillants", 
                    "Category": "Menace interne",
                    "MotivationResources": "Vengeance ou gain personnel - Accès privilégié",
                    "Targeting": "Données confidentielles et systèmes internes"
                },
                {
                    "Source_ID": "RS004",
                    "Label": "Hacktivistes",
                    "Category": "Activisme numérique",
                    "MotivationResources": "Idéologie - Outils collaboratifs",
                    "Targeting": "Sites web et communication publique"
                },
                {
                    "Source_ID": "RS005",
                    "Label": "Prestataires compromis",
                    "Category": "Chaîne d'approvisionnement",
                    "MotivationResources": "Accès indirect - Privilèges étendus",
                    "Targeting": "Systèmes clients via relations de confiance"
                }
            ],
            
            # Catalogue des scénarios stratégiques
            "tbl_Scenario": [
                {
                    "Scenario_ID": "SR001",
                    "Risk_Source": "RS001",
                    "Target_Objective": "Vol de données clients",
                    "Attack_Path": "Attaque externe ciblée",
                    "Motivation": "Revente de données personnelles"
                },
                {
                    "Scenario_ID": "SR002",
                    "Risk_Source": "RS003", 
                    "Target_Objective": "Sabotage système",
                    "Attack_Path": "Abus de privilèges internes",
                    "Motivation": "Vengeance après licenciement"
                },
                {
                    "Scenario_ID": "SR003",
                    "Risk_Source": "RS002",
                    "Target_Objective": "Espionnage industriel",
                    "Attack_Path": "APT ciblée longue durée",
                    "Motivation": "Avantage concurrentiel étatique"
                },
                {
                    "Scenario_ID": "SR004",
                    "Risk_Source": "RS004",
                    "Target_Objective": "Défiguration site web",
                    "Attack_Path": "Attaque de surface publique",
                    "Motivation": "Message politique ou social"
                }
            ],
            
            # Catalogue des scénarios opérationnels (OV)
            "tbl_OV": [
                {
                    "OV_ID": "OV001",
                    "Strategic_Scenario": "SR001",
                    "Attack_Vector": "Phishing et ingénierie sociale",
                    "Operational_Steps": "Reconnaissance > Intrusion > Persistance > Exfiltration"
                },
                {
                    "OV_ID": "OV002",
                    "Strategic_Scenario": "SR002",
                    "Attack_Vector": "Accès physique et logique",
                    "Operational_Steps": "Planification > Exécution > Effacement traces"
                },
                {
                    "OV_ID": "OV003",
                    "Strategic_Scenario": "SR003",
                    "Attack_Vector": "Compromission chaîne logicielle",
                    "Operational_Steps": "Infiltration > Installation > C&C > Collecte > Exfiltration"
                },
                {
                    "OV_ID": "OV004",
                    "Strategic_Scenario": "SR004",
                    "Attack_Vector": "Exploitation vulnérabilités web",
                    "Operational_Steps": "Scan > Exploitation > Défiguration > Revendication"
                }
            ],
            
            # Table des types d'actifs avec libellés complets
            "tbl_AssetType": [
                {"Asset_Type_ID": "AT001", "Libelle": "Serveur", "Description": "Serveurs physiques et virtuels"},
                {"Asset_Type_ID": "AT002", "Libelle": "Base de données", "Description": "Systèmes de gestion de base de données"},
                {"Asset_Type_ID": "AT003", "Libelle": "Application", "Description": "Applications métier et logiciels"},
                {"Asset_Type_ID": "AT004", "Libelle": "Réseau", "Description": "Infrastructure réseau et télécoms"},
                {"Asset_Type_ID": "AT005", "Libelle": "Poste de travail", "Description": "Postes utilisateurs et périphériques"},
                {"Asset_Type_ID": "AT006", "Libelle": "Données", "Description": "Données et informations sensibles"},
                {"Asset_Type_ID": "AT007", "Libelle": "Personnel", "Description": "Ressources humaines et compétences"},
                {"Asset_Type_ID": "AT008", "Libelle": "Locaux", "Description": "Sites et infrastructures physiques"},
                {"Asset_Type_ID": "AT009", "Libelle": "Processus", "Description": "Processus métier et procédures"}
            ],
            
            # Table des parties prenantes avec libellés complets
            "tbl_Stakeholder": [
                {"Stakeholder_ID": "SH001", "Libelle": "DSI", "Description": "Direction des Systèmes d'Information"},
                {"Stakeholder_ID": "SH002", "Libelle": "Direction", "Description": "Direction Générale"},
                {"Stakeholder_ID": "SH003", "Libelle": "RSSI", "Description": "Responsable Sécurité des Systèmes d'Information"},
                {"Stakeholder_ID": "SH004", "Libelle": "DPO", "Description": "Délégué à la Protection des Données"},
                {"Stakeholder_ID": "SH005", "Libelle": "Métier", "Description": "Directions métier"},
                {"Stakeholder_ID": "SH006", "Libelle": "Support", "Description": "Support technique et maintenance"},
                {"Stakeholder_ID": "SH007", "Libelle": "Externe", "Description": "Prestataires externes"},
                {"Stakeholder_ID": "SH008", "Libelle": "Fournisseur", "Description": "Fournisseurs et partenaires"}
            ]
        }

    def generate_template(self, output_path: Path, pme_profile: bool = False) -> None:
        """Génère le template Excel complet conforme EBIOS RM."""
        logger.info("Génération du template Excel EBIOS RM...")
        
        if pme_profile:
            logger.info("Mode PME/TPE activé - Configuration simplifiée")
        
        # Supprimer la feuille par défaut
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
        
        # 1. Créer l'onglet de références avec toutes les tables
        self._create_references_sheet()
        
        # 2. Créer l'onglet de configuration EBIOS RM
        self._create_config_sheet(pme_profile)
        
        # 3. Créer les onglets de travail EBIOS RM complets
        self._create_atelier1_socle()
        self._create_atelier2_sources()
        self._create_atelier3_scenarios() 
        self._create_atelier4_operationnels()
        self._create_atelier5_traitement()
        
        # **CORRECTION** : Créer la table Incidents pour les KPI
        self._create_incidents_table()
        
        self._create_synthese()
        
        # 4. Configuration finale
        self.wb["__REFS"].sheet_state = "veryHidden"  # Masquer l'onglet références
        self.wb.active = self.wb["Config_EBIOS"]    # Définir la feuille de config comme active
        
        # 5. Sauvegarder le classeur
        self.wb.save(output_path)
        logger.info(f"Template généré avec succès : {output_path}")

    def _create_incidents_table(self) -> None:
        """Crée la table Incidents pour alimenter les KPI Velocity."""
        ws = self.wb.create_sheet("Incidents")
        
        # En-têtes de la table Incidents
        headers = ["ID", "Date_Detection", "Date_Reponse", "Temps_Detection", "Temps_Reponse", "Temps_Resolution", "Statut", "Gravite"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Données d'exemple pour les tests
        sample_incidents = [
            ["INC001", "2024-01-15 09:00", "2024-01-15 09:30", 6, 0.5, 48, "Résolu", "Moyen"],
            ["INC002", "2024-01-20 14:00", "2024-01-20 14:15", 12, 0.25, 24, "Résolu", "Élevé"],
            ["INC003", "2024-02-01 08:30", "2024-02-01 09:00", 18, 0.5, 72, "Résolu", "Critique"],
            ["INC004", "2024-02-10 16:00", "2024-02-10 16:45", 24, 0.75, 96, "En cours", "Faible"],
            ["INC005", "2024-02-15 10:15", "2024-02-15 10:30", 8, 0.25, 36, "Résolu", "Moyen"],
            ["INC006", "2024-03-01 13:45", "2024-03-01 14:30", 30, 0.75, 120, "En cours", "Critique"],
            ["INC007", "2024-03-05 11:20", "2024-03-05 11:35", 15, 0.25, 48, "Résolu", "Élevé"],
            ["INC008", "2024-03-10 07:30", "2024-03-10 08:00", 20, 0.5, 60, "Résolu", "Moyen"]
        ]
        
        for row_idx, row_data in enumerate(sample_incidents, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Créer la Table Excel
        end_row = len(sample_incidents) + 1
        table_ref = f"A1:H{end_row}"
        
        table = Table(displayName="Incidents", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showRowStripes=True
        )
        ws.add_table(table)
        
        # Masquer cette feuille (données techniques)
        ws.sheet_state = "hidden"
        
        logger.info("✅ Table Incidents créée pour les KPI Velocity")

    def _create_references_sheet(self) -> None:
        """Crée l'onglet __REFS avec toutes les tables de référence et plages nommées."""
        ws = self.wb.create_sheet("__REFS")
        
        current_col = 1  # Position de départ pour les tables
        
        # Créer chaque table de référence
        for table_name, data in self.reference_data.items():
            if not data:
                continue
                
            # Extraire les en-têtes de la première ligne de données
            headers = list(data[0].keys())
            start_row = 1
            start_col = current_col
            
            # Écrire les en-têtes avec style
            for i, header in enumerate(headers):
                cell = ws.cell(row=start_row, column=start_col + i, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Écrire les données de référence
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, (key, value) in enumerate(row_data.items()):
                    ws.cell(row=row_idx, column=start_col + col_idx, value=value)
            
            # Créer la table Excel pour cette référence
            end_row = len(data) + 1
            end_col = start_col + len(headers) - 1
            table_ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
            
            table = Table(displayName=table_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", 
                showFirstColumn=False,
                showRowStripes=True
            )
            ws.add_table(table)
            
            # Créer les plages nommées pour les validations
            self._create_named_ranges_for_table(ws, table_name, headers, start_col, start_row, end_row)
            
            # Passer à la position suivante (avec espacement)
            current_col = end_col + 2
        
        # **CORRECTION 1.1** : Ajouter les plages pour Pertinence/Exposition
        pertinence_range = f"__REFS!$A$2:$A$4"  # Ajuster selon position réelle
        exposition_range = f"__REFS!$A$2:$A$4"   # Ajuster selon position réelle
        
        self.wb.defined_names["Pertinence"] = DefinedName("Pertinence", attr_text=pertinence_range)
        self.wb.defined_names["Exposition"] = DefinedName("Exposition", attr_text=exposition_range)

    def _create_named_ranges_for_table(self, ws, table_name: str, headers: List[str], 
                                      start_col: int, start_row: int, end_row: int) -> None:
        """Crée les plages nommées nécessaires pour les validations et formules XLOOKUP."""
        
        # **CORRECTION COMPLÈTE** : Mapping avec toutes les plages requises pour validation
        range_mappings = {
            "tbl_Gravite": {
                "Libelle": "Gravite", 
                "ID": "tbl_Gravite_ID", 
                "Valeur": "tbl_Gravite_Valeur"
            },
            "tbl_Vraisemblance": {
                "Libelle": "Vraisemblance", 
                "ID": "tbl_Vraisemblance_ID", 
                "Valeur": "tbl_Vraisemblance_Valeur"
            }, 
            "tbl_ValeurMetier": {
                "ID": "Valeur_Metier", 
                "Libelle": "tbl_ValeurMetier_Libelle", 
                "Valeur": "tbl_ValeurMetier_Valeur"
            },
            "tbl_Pertinence": {
                "Libelle": "Pertinence", 
                "ID": "tbl_Pertinence_ID", 
                "Valeur": "tbl_Pertinence_Valeur"
            },
            "tbl_Exposition": {
                "Libelle": "Exposition", 
                "ID": "tbl_Exposition_ID", 
                "Valeur": "tbl_Exposition_Valeur"
            },
            "tbl_Measure": {
                "Measure_ID": "Measure_ID", 
                "Libelle": "tbl_Measure_Label",
                "Category": "tbl_Measure_Category",
                "Cout": "tbl_Measure_Cout",
                "Efficacite_pct": "tbl_Measure_Efficacite",
                "AnnexA_Control": "tbl_Measure_AnnexA"
            },
            "tbl_Source": {"Source_ID": "Source_ID"},
            "tbl_Scenario": {"Scenario_ID": "Scenario_ID"},
            "tbl_OV": {"OV_ID": "OV_ID"},
            "tbl_AssetType": {"Asset_Type_ID": "Asset_Type", "Libelle": "tbl_AssetType_Label"},
            "tbl_Stakeholder": {"Stakeholder_ID": "Stakeholder_ID", "Libelle": "tbl_Stakeholder_Label"}
        }
        
        # Créer les plages nommées principales avec vérification
        if table_name in range_mappings:
            for header_name, range_name in range_mappings[table_name].items():
                if header_name in headers:
                    col_idx = headers.index(header_name)
                    col_letter = get_column_letter(start_col + col_idx)
                    
                    # **CORRECTION** : Vérifier que la plage est valide
                    if end_row > start_row:
                        range_ref = f"__REFS!${col_letter}$2:${col_letter}${end_row}"
                        defined_name = DefinedName(range_name, attr_text=range_ref)
                        
                        # Ajouter la plage nommée au classeur
                        self.wb.defined_names[range_name] = defined_name
                        logger.info(f"✅ Plage nommée créée: {range_name} = {range_ref}")
        
        # **CORRECTION** : Créer des plages détaillées pour XLOOKUP
        detailed_mappings = {
            "tbl_Source": ["Source_ID", "Label", "Category", "MotivationResources", "Targeting"],
            "tbl_Scenario": ["Scenario_ID", "Risk_Source", "Target_Objective", "Attack_Path", "Motivation"],
            "tbl_OV": ["OV_ID", "Strategic_Scenario", "Attack_Vector", "Operational_Steps"]
        }
        
        if table_name in detailed_mappings:
            # Créer des plages individuelles pour chaque colonne
            for header_name in detailed_mappings[table_name]:
                if header_name in headers:
                    col_idx = headers.index(header_name)
                    col_letter = get_column_letter(start_col + col_idx)
                    
                    range_name = f"{table_name}_{header_name}"
                    
                    if end_row > start_row:
                        range_ref = f"__REFS!${col_letter}$2:${col_letter}${end_row}"
                        defined_name = DefinedName(range_name, attr_text=range_ref)
                        
                        self.wb.defined_names[range_name] = defined_name
                        logger.info(f"✅ Plage détaillée créée: {range_name} = {range_ref}")

    def _create_atelier1_socle(self) -> None:
        """Crée l'Atelier 1 - Socle avec les validations appropriées et listes dépendantes."""
        ws = self.wb.create_sheet("Atelier1_Socle")
        
        # **CORRECTION 3** : En-têtes étendus avec sous-type
        headers = [
            "ID_Actif", "Type", "Sous_Type", "Libellé", "Description", "Gravité",
            "Confidentialité", "Intégrité", "Disponibilité", 
            "Valeur_Métier", "Propriétaire", "Score_Risque"
        ]
        
        # Créer les en-têtes avec style
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Appliquer les validations de données selon EBIOS RM
        validations_config = {
            2: "Asset_Type",        # Colonne Type
            5: "Gravite",           # Colonne Gravité
            6: "Gravite",           # Colonne Confidentialité  
            7: "Gravite",           # Colonne Intégrité
            8: "Gravite",           # Colonne Disponibilité
            9: "Valeur_Metier",     # Colonne Valeur Métier
            10: "Stakeholder_ID",   # Colonne Propriétaire
        }
        
        # Appliquer chaque validation avec le signe "=" obligatoire
        for col_num, range_name in validations_config.items():
            dv = DataValidation(type="list", formula1=f"={range_name}", allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(col_num)}2:{get_column_letter(col_num)}100")
        
        # **CORRECTION 3** : Formule de risque pondérée corrigée avec XLOOKUP
        for row in range(2, 101):
            risk_formula = f"""=IF(AND(E{row}<>"",F{row}<>"",G{row}<>"",H{row}<>"",I{row}<>""),
XLOOKUP(E{row},Gravite,tbl_Gravite_ID)*
XLOOKUP(F{row},Gravite,tbl_Gravite_ID)*
XLOOKUP(G{row},Gravite,tbl_Gravite_ID)*
XLOOKUP(H{row},Gravite,tbl_Gravite_ID)*
XLOOKUP(I{row},Valeur_Metier,tbl_ValeurMetier_ID),"")"""
            
            cell = ws.cell(row=row, column=11, value=risk_formula)
            self._format_formula_cell(cell)
        
        # Ajouter des exemples de données
        sample_data = [
            ["A001", "", "Base clients", "Base de données des clients", "", "", "", "", "", "", ""],
            ["A002", "", "Serveur web", "Serveur d'application web", "", "", "", "", "", "", ""],
            ["A003", "", "Plans stratégiques", "Documents confidentiels", "", "", "", "", "", "", ""],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                if col_idx != 11:  # Ne pas écraser la formule de score
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Appliquer la protection et le formatage
        self._apply_sheet_protection(ws)
        ws.freeze_panes = "B2"  # Figer les volets

    def _create_atelier2_sources(self) -> None:
        """Crée l'Atelier 2 - Sources de risque avec formules XLOOKUP fonctionnelles."""
        ws = self.wb.create_sheet("Atelier2_Sources") 
        
        # En-têtes français selon EBIOS RM Atelier 2
        headers = [
            "ID_Source", "Libellé", "Catégorie", "Motivation_Ressources", 
            "Ciblage", "Pertinence", "Exposition", "Commentaires"
        ]
        
        # Créer les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Validation ID_Source avec plage nommée
        dv = DataValidation(type="list", formula1="=Source_ID", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("A2:A1000")
        
        # Formules XLOOKUP corrigées pour recherche dans toute la table
        for row in range(2, 101):
            # Libellé - recherche dans la table complète
            cell = ws.cell(row=row, column=2, value=f"=IFERROR(INDEX(tbl_Source_Label,MATCH(A{row},tbl_Source_Source_ID,0)),\"\")")
            self._format_formula_cell(cell)
            
            # Catégorie
            cell = ws.cell(row=row, column=3, value=f"=IFERROR(INDEX(tbl_Source_Category,MATCH(A{row},tbl_Source_Source_ID,0)),\"\")")
            self._format_formula_cell(cell)
            
            # Motivation_Ressources
            cell = ws.cell(row=row, column=4, value=f"=IFERROR(INDEX(tbl_Source_MotivationResources,MATCH(A{row},tbl_Source_Source_ID,0)),\"\")")
            self._format_formula_cell(cell)
            
            # Ciblage
            cell = ws.cell(row=row, column=5, value=f"=IFERROR(INDEX(tbl_Source_Targeting,MATCH(A{row},tbl_Source_Source_ID,0)),\"\")")
            self._format_formula_cell(cell)
        
        # Validations pour les niveaux d'évaluation
        for col in [6, 7]:  # Pertinence, Exposition
            dv = DataValidation(type="list", formula1="=Gravite", allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}1000")
        
        # **CORRECTION 2.1** : Validations dédiées Pertinence/Exposition avec messages
        pertinence_dv = DataValidation(type="list", formula1="=Pertinence", allow_blank=True)
        pertinence_dv.error = "Niveau de pertinence invalide. Choisissez : Faible, Modérée ou Forte"
        pertinence_dv.errorTitle = "Erreur de saisie - Pertinence"
        pertinence_dv.prompt = "Évaluez le niveau de pertinence de cette source pour votre organisation (échelle 1-3)"
        pertinence_dv.promptTitle = "Guide d'évaluation - Pertinence"
        pertinence_dv.showErrorMessage = True
        pertinence_dv.showInputMessage = True
        pertinence_dv.showDropDown = True  # **CORRECTION 6** : Flèche visible
        ws.add_data_validation(pertinence_dv)
        pertinence_dv.add("F2:F1000")
        
        exposition_dv = DataValidation(type="list", formula1="=Exposition", allow_blank=True)
        exposition_dv.error = "Niveau d'exposition invalide. Choisissez : Limitée, Significative ou Maximale"
        exposition_dv.errorTitle = "Erreur de saisie - Exposition"
        exposition_dv.prompt = "Évaluez le niveau d'exposition de votre organisation à cette source (échelle 1-3)"
        exposition_dv.promptTitle = "Guide d'évaluation - Exposition"
        exposition_dv.showErrorMessage = True
        exposition_dv.showInputMessage = True
        exposition_dv.showDropDown = True
        ws.add_data_validation(exposition_dv)
        exposition_dv.add("G2:G1000")
        
        # Protection et formatage
        self._apply_sheet_protection(ws)
        ws.freeze_panes = "B2"

    def _create_atelier3_scenarios(self) -> None:
        """Crée l'Atelier 3 - Scénarios stratégiques avec formules XLOOKUP et calcul de risque pondéré."""
        ws = self.wb.create_sheet("Atelier3_Scenarios")
        
        # En-têtes français selon EBIOS RM Atelier 3
        headers = [
            "ID_Scénario", "Source_Risque", "Objectif_Visé", "Chemin_Attaque",
            "Motivation", "Gravité", "Vraisemblance", "Valeur_Métier", "Risque_Calculé"
        ]
        
        # Créer les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Validation ID_Scénario
        dv = DataValidation(type="list", formula1="=Scenario_ID", allow_blank=True)
        dv.showDropDown = False  # Force l'affichage de la flèche
        ws.add_data_validation(dv)
        dv.add("A2:A1000")
        
        # Formules XLOOKUP pour remplissage automatique
        for row in range(2, 101):
            # Source_Risque - formule XLOOKUP corrigée
            cell = ws.cell(row=row, column=2, value=f"=IFERROR(XLOOKUP(A{row},tbl_Scenario_Scenario_ID,tbl_Scenario_Risk_Source),\"\")")
            self._format_formula_cell(cell)
            
            # Objectif_Visé
            cell = ws.cell(row=row, column=3, value=f"=IFERROR(XLOOKUP(A{row},tbl_Scenario_Scenario_ID,tbl_Scenario_Target_Objective),\"\")")
            self._format_formula_cell(cell)
            
            # Chemin_Attaque
            cell = ws.cell(row=row, column=4, value=f"=IFERROR(XLOOKUP(A{row},tbl_Scenario_Scenario_ID,tbl_Scenario_Attack_Path),\"\")")
            self._format_formula_cell(cell)
            
            # Motivation
            cell = ws.cell(row=row, column=5, value=f"=IFERROR(XLOOKUP(A{row},tbl_Scenario_Scenario_ID,tbl_Scenario_Motivation),\"\")")
            self._format_formula_cell(cell)
            
            # **CORRECTION** : Calcul du risque pondéré avec XLOOKUP
            risk_formula = f"""=IF(AND(F{row}<>"",G{row}<>"",H{row}<>""),
XLOOKUP(F{row},Gravite,tbl_Gravite_Valeur)*
XLOOKUP(G{row},Vraisemblance,tbl_Vraisemblance_Valeur)*
XLOOKUP(H{row},Valeur_Metier,tbl_ValeurMetier_Valeur),"")"""
            
            cell = ws.cell(row=row, column=9, value=risk_formula)
            self._format_formula_cell(cell)
        
        # **CORRECTION COMPLÈTE** : Validations avec flèches visibles et messages d'erreur
        validation_configs = [
            {
                "column": "F",  # Gravité
                "range": "F2:F1000",
                "formula": "=Gravite",
                "error": "Niveau de gravité invalide. Choisissez : Négligeable, Limité, Important ou Critique",
                "prompt": "Sélectionnez le niveau de gravité du scénario (échelle 1-4)"
            },
            {
                "column": "G",  # Vraisemblance
                "range": "G2:G1000", 
                "formula": "=Vraisemblance",
                "error": "Niveau de vraisemblance invalide. Choisissez : Minimal, Significatif, Élevé ou Maximal",
                "prompt": "Sélectionnez le niveau de vraisemblance du scénario (échelle 1-4)"
            },
            {
                "column": "H",  # Valeur Métier
                "range": "H2:H1000",
                "formula": "=Valeur_Metier", 
                "error": "Valeur métier invalide. Choisissez un niveau entre 1 et 15",
                "prompt": "Sélectionnez la valeur métier de l'actif ciblé (échelle 1-15)"
            }
        ]
        
        for config in validation_configs:
            dv = DataValidation(type="list", formula1=config["formula"], allow_blank=True)
            dv.showDropDown = False  # **CORRECTION** : Force l'affichage de la flèche
            dv.error = config["error"]
            dv.errorTitle = "Erreur de validation - Atelier 3"
            dv.prompt = config["prompt"] 
            dv.promptTitle = "Guide de saisie"
            dv.showErrorMessage = True
            dv.showInputMessage = True
            ws.add_data_validation(dv)
            dv.add(config["range"])
        
        # Protection et formatage
        self._apply_sheet_protection(ws)
        ws.freeze_panes = "B2"

    def _create_atelier4_operationnels(self) -> None:
        """Crée l'Atelier 4 - Scénarios opérationnels avec calculs automatiques."""
        ws = self.wb.create_sheet("Atelier4_Operationnels")
        
        # **CORRECTION** : En-têtes étendus avec risque résiduel
        headers = [
            "ID_OV", "Scénario_Stratégique", "Vecteur_Attaque", "Étapes_Opérationnelles",
            "Contrôles_Existants", "Vraisemblance_Résiduelle", "Impact", "Risque_Initial", 
            "Mesures_Appliquées", "Efficacité_Totale", "Risque_Résiduel", "Niveau_Risque_Final"
        ]
        
        # Créer les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Validation ID_OV avec flèche visible
        dv = DataValidation(type="list", formula1="=OV_ID", allow_blank=True)
        dv.showDropDown = False
        ws.add_data_validation(dv)
        dv.add("A2:A1000")
        
        # **CORRECTION** : Formules XLOOKUP pour auto-complétion
        for row in range(2, 101):
            # Scénario_Stratégique - formule XLOOKUP corrigée
            cell = ws.cell(row=row, column=2, value=f"=IFERROR(XLOOKUP(A{row},tbl_OV_OV_ID,tbl_OV_Strategic_Scenario),\"\")")
            self._format_formula_cell(cell)
            
            # Vecteur_Attaque
            cell = ws.cell(row=row, column=3, value=f"=IFERROR(XLOOKUP(A{row},tbl_OV_OV_ID,tbl_OV_Attack_Vector),\"\")")
            self._format_formula_cell(cell)
            
            # Étapes_Opérationnelles
            cell = ws.cell(row=row, column=4, value=f"=IFERROR(XLOOKUP(A{row},tbl_OV_OV_ID,tbl_OV_Operational_Steps),\"\")")
            self._format_formula_cell(cell)
            
            # **CORRECTION** : Calcul automatique du risque initial
            risk_initial_formula = f"""=IF(AND(F{row}<>"",G{row}<>""),
XLOOKUP(F{row},Vraisemblance,tbl_Vraisemblance_Valeur)*
XLOOKUP(G{row},Gravite,tbl_Gravite_Valeur),"")"""
            
            cell = ws.cell(row=row, column=8, value=risk_initial_formula)
            self._format_formula_cell(cell)
            
            # **CORRECTION** : Efficacité totale depuis les mesures
            efficacite_formula = f"=IFERROR(XLOOKUP(I{row},Measure_ID,tbl_Measure_Efficacite),\"\")"
            cell = ws.cell(row=row, column=10, value=efficacite_formula)
            self._format_formula_cell(cell)
            
            # **CORRECTION** : Calcul du risque résiduel avec efficacité des mesures
            risque_residuel_formula = f"""=IF(AND(ISNUMBER(H{row}),ISNUMBER(J{row})),
H{row}*(1-J{row}/100),"")"""
            
            cell = ws.cell(row=row, column=11, value=risque_residuel_formula)
            self._format_formula_cell(cell)
            
            # **CORRECTION** : Niveau de risque final automatique
            niveau_final_formula = f"""=IF(K{row}<>"",
IF(K{row}>=12,"Critique",
IF(K{row}>=8,"Élevé",
IF(K{row}>=4,"Moyen","Faible"))),"")"""
            
            cell = ws.cell(row=row, column=12, value=niveau_final_formula)
            self._format_formula_cell(cell)
        
        # **CORRECTION COMPLÈTE** : Validations avec flèches visibles
        validation_configs = [
            {
                "column": "F",  # Vraisemblance_Résiduelle
                "range": "F2:F1000",
                "formula": "=Vraisemblance",
                "error": "Niveau de vraisemblance invalide",
                "prompt": "Évaluez la vraisemblance résiduelle après contrôles existants"
            },
            {
                "column": "G",  # Impact
                "range": "G2:G1000",
                "formula": "=Gravite", 
                "error": "Niveau d'impact invalide",
                "prompt": "Évaluez l'impact potentiel du scénario opérationnel"
            },
            {
                "column": "I",  # Mesures_Appliquées
                "range": "I2:I1000",
                "formula": "=Measure_ID",
                "error": "Cette mesure n'existe pas dans le catalogue ISO 27001",
                "prompt": "Sélectionnez une mesure de sécurité du catalogue Annex A"
            }
        ]
        
        for config in validation_configs:
            dv = DataValidation(type="list", formula1=config["formula"], allow_blank=True)
            dv.showDropDown = False  # **CORRECTION** : Force l'affichage de la flèche
            dv.error = config["error"]
            dv.errorTitle = "Erreur de validation - Atelier 4"
            dv.prompt = config["prompt"]
            dv.promptTitle = "Guide de saisie"
            dv.showErrorMessage = True
            dv.showInputMessage = True
            ws.add_data_validation(dv)
            dv.add(config["range"])
        
        # Protection et formatage
        self._apply_sheet_protection(ws)
        ws.freeze_panes = "B2"

    def _create_atelier5_traitement(self) -> None:
        """Crée l'Atelier 5 - Traitement du risque avec plan d'action détaillé."""
        ws = self.wb.create_sheet("Atelier5_Traitement")
        
        # **CORRECTION** : En-têtes étendus avec risque résiduel
        headers = [
            "ID_Risque", "Scénario_Lié", "Niveau_Initial", "Option_Traitement", 
            "Mesure_Choisie", "Contrôle_AnnexA", "Responsable", "Échéance", 
            "Coût_Estimé", "Efficacité_Attendue", "Niveau_Résiduel", "Statut_Mise_en_Œuvre"
        ]
        
        # Créer les en-têtes avec style
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Options de traitement selon EBIOS RM
        options_traitement = ["Réduire", "Éviter", "Transférer", "Accepter"]
        statuts = ["Planifiée", "En cours", "Terminée", "Reportée", "Annulée"]
        
        # **CORRECTION** : Formules automatiques avec auto-complétion des mesures
        for row in range(2, 51):
            # Contrôle Annex A automatique depuis la mesure
            cell = ws.cell(row=row, column=6, value=f"=IFERROR(XLOOKUP(E{row},Measure_ID,tbl_Measure_AnnexA),\"\")")
            self._format_formula_cell(cell)
            
            # Coût estimé depuis catalogue
            cell = ws.cell(row=row, column=9, value=f"=IFERROR(XLOOKUP(E{row},Measure_ID,tbl_Measure_Cout),\"\")")
            self._format_formula_cell(cell)
            
            # **CORRECTION** : Efficacité attendue depuis catalogue avec XLOOKUP
            cell = ws.cell(row=row, column=10, value=f"=IFERROR(XLOOKUP(E{row},Measure_ID,tbl_Measure_Efficacite),\"\")")
            self._format_formula_cell(cell)
            
            # **CORRECTION** : Calcul du risque résiduel
            risque_residuel_formula = f"""=IF(AND(ISNUMBER(C{row}),ISNUMBER(J{row})),
C{row}*(1-J{row}/100),"")"""
            cell = ws.cell(row=row, column=11, value=risque_residuel_formula)
            self._format_formula_cell(cell)
        
        # **CORRECTION** : Validations de données avec flèches visibles
        validation_configs = [
            {
                "column": 3,  # Niveau_Initial
                "source": "=Gravite",
                "description": "Niveau initial de risque"
            },
            {
                "column": 4,  # Option_Traitement  
                "source": options_traitement,
                "description": "Option de traitement du risque"
            },
            {
                "column": 5,  # Mesure_Choisie
                "source": "=Measure_ID",
                "description": "Mesure de sécurité du catalogue ISO 27001"
            },
            {
                "column": 7,  # Responsable
                "source": "=Stakeholder_ID",
                "description": "Responsable de la mise en œuvre"
            },
            {
                "column": 12,  # Statut_Mise_en_Œuvre
                "source": statuts,
                "description": "Statut d'avancement de la mesure"
            }
        ]
        
        for config in validation_configs:
            if isinstance(config["source"], list):
                dv = DataValidation(type="list", formula1=f'"{",".join(config["source"])}"', allow_blank=True)
            else:
                dv = DataValidation(type="list", formula1=config["source"], allow_blank=True)
            
            # **CORRECTION** : Configuration complète avec flèches visibles
            dv.showDropDown = False  # Force l'affichage de la flèche
            dv.prompt = f"Sélectionnez {config['description']}"
            dv.promptTitle = "Aide à la saisie"
            dv.error = f"Valeur non autorisée pour {config['description']}"
            dv.errorTitle = "Erreur de validation"
            dv.showErrorMessage = True
            dv.showInputMessage = True
            
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(config['column'])}2:{get_column_letter(config['column'])}100")
        
        # Appliquer la protection et le formatage
        self._apply_sheet_protection(ws)
        ws.freeze_panes = "B2"

    def _create_heatmap_visualization(self) -> None:
        """Crée l'onglet Heat-map avec matrice de risque visuelle."""
        ws = self.wb.create_sheet("HeatMap_Risques")
        
        # Titre principal
        ws.merge_cells("A1:J1")
        title = ws["A1"]
        title.value = "🔥 MATRICE DE CHALEUR - CARTOGRAPHIE DES RISQUES"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        
        # **INNOVATION** : Matrice 4×4 avec mise en forme conditionnelle
        # Headers Vraisemblance (colonnes)
        likelihood_labels = ["", "Minimal", "Significatif", "Élevé", "Maximal"]
        for col, label in enumerate(likelihood_labels, 2):
            cell = ws.cell(row=3, column=col, value=label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Headers Gravité (lignes) 
        gravity_labels = ["Négligeable", "Limité", "Important", "Critique"]
        risk_matrix_values = [
            [1, 2, 3, 4],      # Négligeable
            [2, 4, 6, 8],      # Limité  
            [3, 6, 9, 12],     # Important
            [4, 8, 12, 16]     # Critique
        ]
        
        risk_colors = {
            (1, 2, 3): "27AE60",     # Vert - Faible
            (4, 6): "F39C12",        # Orange - Moyen  
            (8, 9): "E74C3C",        # Rouge - Élevé
            (12, 16): "C0392B"       # Rouge foncé - Critique
        }
        
        for row_idx, (gravity_label, risk_row) in enumerate(zip(gravity_labels, risk_matrix_values), 4):
            # Label gravité
            gravity_cell = ws.cell(row=row_idx, column=1, value=gravity_label)
            gravity_cell.font = Font(bold=True, color="FFFFFF")
            gravity_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            gravity_cell.alignment = Alignment(horizontal="center")
            
            # Valeurs de risque avec couleurs
            for col_idx, risk_value in enumerate(risk_row, 2):
                risk_cell = ws.cell(row=row_idx, column=col_idx, value=risk_value)
                risk_cell.alignment = Alignment(horizontal="center", vertical="center")
                risk_cell.font = Font(bold=True, size=14, color="FFFFFF")
                
                # Appliquer couleur selon valeur
                for values, color in risk_colors.items():
                    if risk_value in values:
                        risk_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        break
        
        # **INNOVATION** : Tableau de répartition dynamique des scénarios
        ws["A10"] = "📊 RÉPARTITION DES SCÉNARIOS PAR ZONE DE RISQUE"
        ws["A10"].font = Font(size=12, bold=True)
        
        distribution_headers = ["Zone de Risque", "Nombre Scénarios", "% Total", "Actions Recommandées"]
        for col, header in enumerate(distribution_headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
        
        zones_risk = [
            ("🟢 Acceptable (1-3)", '=COUNTIFS(Atelier4_Operationnels[Risque_Residuel],">=1",Atelier4_Operationnels[Risque_Residuel],"<=3")', "Surveillance"),
            ("🟡 Tolérable (4-6)", '=COUNTIFS(Atelier4_Operationnels[Risque_Residuel],">=4",Atelier4_Operationnels[Risque_Residuel],"<=6")', "Mesures ciblées"),
            ("🟠 Inacceptable (8-9)", '=COUNTIFS(Atelier4_Operationnels[Risque_Residuel],">=8",Atelier4_Operationnels[Risque_Residuel],"<=9")', "Plan d'action immédiat"),
            ("🔴 Critique (12-16)", '=COUNTIFS(Atelier4_Operationnels[Risque_Residuel],">=12",Atelier4_Operationnels[Risque_Residuel],"<=16")', "Traitement d'urgence")
        ]
        
        for row_idx, (zone, formula, action) in enumerate(zones_risk, 12):
            ws.cell(row=row_idx, column=1, value=zone)
            ws.cell(row=row_idx, column=2, value=formula)
            ws.cell(row=row_idx, column=3, value=f'=IF(SUM(B12:B15)>0,B{row_idx}/SUM(B12:B15)*100,0)&"%"')
            ws.cell(row=row_idx, column=4, value=action)

    def _create_synthese(self) -> None:
        """Crée l'onglet Synthèse avec indicateurs clés et KPI avancés."""
        ws = self.wb.create_sheet("Synthese")
        
        # Titre
        ws["A1"] = "📊 SYNTHÈSE EXÉCUTIVE - ANALYSE DES RISQUES"
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        ws.merge_cells("A1:F1")
        
        # **CORRECTION 3.2** : Section KPI Velocity/Preparedness
        ws["A3"] = "⚡ INDICATEURS VELOCITY"
        ws["A3"].font = Font(size=12, bold=True, color="3498DB")
        
        velocity_headers = ["KPI", "Valeur Actuelle", "Cible", "Performance", "Tendance"]
        for col, header in enumerate(velocity_headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # KPI Velocity - FIX: Corriger les formules avec guillemets échappés
        velocity_kpis = [
            ("Temps détection (h)", "=AVERAGE(Incidents[Temps_Detection])", "24", '=IF(B5<=C5,"✅ Conforme","⚠️ À améliorer")'),
            ("Temps réponse (h)", "=AVERAGE(Incidents[Temps_Reponse])", "4", '=IF(B6<=C6,"✅ Conforme","❌ Non conforme")'),
            ("% résolution < 72h", '=COUNTIFS(Incidents[Temps_Resolution],"<72")/COUNTA(Incidents[ID])*100', "90", '=IF(B7>=C7,"✅ Conforme","⚠️ À améliorer")')
        ]
        
        for row_idx, (kpi_name, formula, target, status_formula) in enumerate(velocity_kpis, 5):
            ws.cell(row=row_idx, column=1, value=kpi_name)
            ws.cell(row=row_idx, column=2, value=formula)
            ws.cell(row=row_idx, column=3, value=target)
            ws.cell(row=row_idx, column=4, value=status_formula)
            ws.cell(row=row_idx, column=5, value="📊")
        
        # **CORRECTION 3.2** : Section Preparedness
        ws["A9"] = "🛡️ INDICATEURS PREPAREDNESS"
        ws["A9"].font = Font(size=12, bold=True, color="27AE60")
        
        for col, header in enumerate(velocity_headers, 1):
            cell = ws.cell(row=10, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        
        preparedness_kpis = [
            ("% actifs couverts", '=COUNTIFS(Atelier1_Socle[Score_Risque],">0")/COUNTA(Atelier1_Socle[ID_Actif])*100', "95"),
            ("% mesures implémentées", '=COUNTIFS(Atelier5_Traitement[Statut_Mise_en_Œuvre],"Terminée")/COUNTA(Atelier5_Traitement[ID_Risque])*100', "80"),
            ("Maturité globale", "=AVERAGE(Atelier4_Operationnels[Niveau_Risque_Final])", "3")
        ]
        
        for row_idx, (kpi_name, formula, target) in enumerate(preparedness_kpis, 11):
            ws.cell(row=row_idx, column=1, value=kpi_name)
            ws.cell(row=row_idx, column=2, value=formula)
            ws.cell(row=row_idx, column=3, value=target)
            ws.cell(row=row_idx, column=4, value=f'=IF(B{row_idx}>=C{row_idx},"✅ Conforme","⚠️ À améliorer")')
            ws.cell(row=row_idx, column=5, value="📈")
        
        # Métriques principales
        ws["A15"] = "🎯 INDICATEURS CLÉS"
        ws["A15"].font = Font(size=12, bold=True)
        
        # En-têtes
        headers = ["Indicateur", "Valeur", "Statut", "Tendance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=16, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.gray_fill
        
        # Données de synthèse avec formules corrigées
        metrics = [
            ["Nombre d'actifs analysés", "=COUNTA(Atelier1_Socle[ID_Actif])", "En cours", "↗️"],
            ["Sources de risque identifiées", "=COUNTA(Atelier2_Sources[ID_Source])", "Complété", "→"],
            ["Scénarios évalués", "=COUNTA(Atelier3_Scenarios[ID_Scénario])", "En cours", "↗️"],
            ["Mesures planifiées", "=COUNTA(Atelier5_Traitement[ID_Risque])", "Planifié", "↗️"]
        ]
        
        for row_idx, metric_data in enumerate(metrics, 17):
            for col_idx, value in enumerate(metric_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        logger.info("✅ Onglet de synthèse créé avec KPI Velocity/Preparedness")

    def _create_kpi_dashboard(self) -> None:
        """Crée l'onglet KPI avec indicateurs Velocity et Preparedness."""
        ws = self.wb.create_sheet("KPI_Dashboard")
        
        # Titre dashboard
        ws.merge_cells("A1:H1")
        title = ws["A1"]
        title.value = "📈 TABLEAU DE BORD - INDICATEURS EBIOS RM"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
        title.alignment = Alignment(horizontal="center")
        
        # **SECTION VELOCITY** : Rapidité de détection et réponse
        ws["A3"] = "⚡ VELOCITY - Rapidité d'intervention"
        ws["A3"].font = Font(size=14, bold=True, color="2C3E50")
        
        velocity_headers = ["Indicateur", "Valeur Actuelle", "Cible", "Statut", "Tendance"]
        for col, header in enumerate(velocity_headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        
        # FIX: Corriger les formules KPI avec échappement approprié
        velocity_kpis = [
            ("Temps détection incident (h)", "=AVERAGE(Incidents[Temps_Detection])", "24", '=IF(B5<=C5,"✅ Conforme","⚠️ À améliorer")'),
            ("Temps réponse incident (h)", "=AVERAGE(Incidents[Temps_Reponse])", "4", '=IF(B6<=C6,"✅ Conforme","❌ Non conforme")'),
            ("% incidents résolus < 72h", '=COUNTIFS(Incidents[Temps_Resolution],"<72")/COUNTA(Incidents[ID])*100', "90", '=IF(B7>=C7,"✅ Conforme","⚠️ À améliorer")')
        ]
        
        for row_idx, (kpi_name, formula, target, status_formula) in enumerate(velocity_kpis, 5):
            ws.cell(row=row_idx, column=1, value=kpi_name)
            ws.cell(row=row_idx, column=2, value=formula)
            ws.cell(row=row_idx, column=3, value=target)
            ws.cell(row=row_idx, column=4, value=status_formula)
            ws.cell(row=row_idx, column=5, value="📊")  # Placeholder pour graphique sparkline
        
        # **SECTION PREPAREDNESS** : Niveau de préparation
        ws["A10"] = "🛡️ PREPAREDNESS - Niveau de préparation"
        ws["A10"].font = Font(size=14, bold=True, color="2C3E50")
        
        for col, header in enumerate(velocity_headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        
        # FIX: Corriger les formules Preparedness
        preparedness_kpis = [
            ("% actifs couverts mesures", '=COUNTIFS(Atelier1_Socle[Score_Risque],">0")/COUNTA(Atelier1_Socle[ID_Actif])*100', "95", '=IF(B12>=C12,"✅ Conforme","⚠️ Exposition")'),
            ("% personnel formé SSI", '=IF(ISERROR(COUNTIFS(Personnel[Formation_SSI],"Oui")/COUNTA(Personnel[ID])*100),75,COUNTIFS(Personnel[Formation_SSI],"Oui")/COUNTA(Personnel[ID])*100)', "90", '=IF(B13>=C13,"✅ Conforme","❌ Formation requise")'),
            ("Maturité globale (1-5)", "=IF(ISERROR(AVERAGE(Maturite[Score_Domaine])),3,AVERAGE(Maturite[Score_Domaine]))", "3", '=IF(B14>=C14,"✅ Mature","⚠️ Amélioration")'),
            ("% mesures implémentées", '=COUNTIFS(Atelier5_Traitement[Statut_Mise_en_Œuvre],"Terminée")/COUNTA(Atelier5_Traitement[ID_Risque])*100', "80", '=IF(B15>=C15,"✅ Conforme","❌ Retard")')
        ]
        
        for row_idx, (kpi_name, formula, target, status_formula) in enumerate(preparedness_kpis, 12):
            ws.cell(row=row_idx, column=1, value=kpi_name)
            ws.cell(row=row_idx, column=2, value=formula)
            ws.cell(row=row_idx, column=3, value=target)
            ws.cell(row=row_idx, column=4, value=status_formula)
            ws.cell(row=row_idx, column=5, value="📈")
        
        # **SECTION SYNTHÈSE** : Vue globale
        ws["A18"] = "🎯 SYNTHÈSE GLOBALE"
        ws["A18"].font = Font(size=14, bold=True, color="2C3E50")
        
        synthesis_formulas = [
            ("Score Global Velocity", "=AVERAGE(B5:B7)"),
            ("Score Global Preparedness", "=AVERAGE(B12:B15)"),
            ("Index Maturité EBIOS", "=(B19+B20)/2"),
            ("Recommandation Prioritaire", '=IF(B21<2.5,"Formation & Outils","Optimisation Continue")')
        ]
        
        for row_idx, (metric, formula) in enumerate(synthesis_formulas, 19):
            ws.cell(row=row_idx, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=formula)

    def _format_formula_cell(self, cell) -> None:
        """Formate une cellule contenant une formule (grise + verrouillée)."""
        # Vérifier que la cellule contient bien une formule
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            cell.fill = self.gray_fill
            cell.protection = Protection(locked=True)
    
    def _apply_sheet_protection(self, ws) -> None:
        """Applique la protection intelligente basée sur les formules."""
        # **CORRECTION 6** : Protection sélective améliorée - ordre recommandé
        for row in ws.iter_rows():
            for cell in row:
                # Déverrouiller par défaut
                cell.protection = Protection(locked=False)
                
                # **CORRECTION 6** : Verrouiller et griser UNIQUEMENT les vraies formules
                if (cell.value and isinstance(cell.value, str) and 
                    cell.value.startswith('=') and cell.data_type == "f"):
                    cell.protection = Protection(locked=True)
                    cell.fill = self.gray_fill
                # Les en-têtes restent verrouillés mais pas grisés
                elif cell.row == 1:
                    cell.protection = Protection(locked=True)
        
        # Activer la protection de la feuille (ordre recommandé par Microsoft)
        ws.protection = SheetProtection(sheet=True, password=None)
        logger.info(f"Protection appliquée sur {ws.title} avec grisage sélectif des formules")

    def _create_config_sheet(self, pme_profile: bool = False) -> None:
        """Crée l'onglet de configuration EBIOS RM."""
        ws = self.wb.create_sheet("Config_EBIOS", 0)  # Première position
        
        # Titre principal
        ws["A1"] = "🔧 CONFIGURATION EBIOS RISK MANAGER"
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:F1")
        ws.row_dimensions[1].height = 25
        
        # Description
        ws["A2"] = "Configuration des paramètres EBIOS RM selon profil organisationnel"
        ws["A2"].font = Font(italic=True, color="7F8C8D")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:F2")
        
        # Section profil
        ws["A4"] = "📋 PROFIL ORGANISATIONNEL"
        ws["A4"].font = Font(size=12, bold=True, color="2C3E50")
        
        profile_text = "PME/TPE - Configuration simplifiée" if pme_profile else "Grande entreprise - Configuration complète"
        ws["A5"] = f"Type d'organisation : {profile_text}"
        ws["A5"].font = Font(size=10)
        
        # Paramètres de configuration
        ws["A7"] = "⚙️ PARAMÈTRES DE CONFIGURATION"
        ws["A7"].font = Font(size=12, bold=True, color="2C3E50")
        
        config_params = [
            ("Échelle de gravité", "4 niveaux (Négligeable, Limité, Important, Critique)"),
            ("Échelle de vraisemblance", "4 niveaux (Minimal, Significatif, Élevé, Maximal)"),
            ("Valeurs métier", "15 niveaux (1-15)"),
            ("Sources de risque", "5 sources principales cataloguées"),
            ("Scénarios stratégiques", "4 scénarios de base configurés"),
            ("Mesures de sécurité", "30 mesures ISO 27001 Annex A")
        ]
        
        for i, (param, value) in enumerate(config_params, 8):
            ws.cell(row=i, column=1, value=f"• {param}").font = Font(bold=True, size=10)
            ws.cell(row=i, column=3, value=value).font = Font(size=10)
        
        # Instructions d'utilisation
        ws["A15"] = "📝 INSTRUCTIONS D'UTILISATION"
        ws["A15"].font = Font(size=12, bold=True, color="2C3E50")
        
        instructions = [
            "1. Renseignez les actifs dans l'Atelier 1 - Socle",
            "2. Analysez les sources de risque dans l'Atelier 2", 
            "3. Définissez les scénarios dans l'Atelier 3",
            "4. Évaluez les mesures dans l'Atelier 4",
            "5. Planifiez le traitement dans l'Atelier 5",
            "6. Consultez la synthèse pour les résultats"
               ]
        
        for i, instruction in enumerate(instructions, 16):
            ws.cell(row=i, column=1, value=instruction).font = Font(size=10)
        
        # Section méthodologie
        ws["A23"] = "📚 MÉTHODOLOGIE EBIOS RISK MANAGER"
        ws["A23"].font = Font(size=12, bold=True, color="2C3E50")
        
        methodology_info = [
            "• Méthode d'analyse des risques SSI de l'ANSSI",
            "• Approche en 5 ateliers pour une analyse complète",
            "• Conformité aux standards ISO 27005 et ISO 31000",
            "• Adaptation aux enjeux de cybersécurité actuels"
        ]
        
        for i, info in enumerate(methodology_info, 24):
            ws.cell(row=i, column=1, value=info).font = Font(size=10, italic=True)
        
        # Formatage des colonnes
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 5
        ws.column_dimensions["C"].width = 50
        
        # Couleur de fond alternée pour la lisibilité
        for row_num in range(8, 14, 2):
            for col_num in range(1, 4):
                ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        logger.info("✅ Onglet de configuration créé")

def main():
    """Point d'entrée principal pour la génération du template EBIOS RM."""
    # Configuration du logging pour avoir des messages visibles
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(),  # Affichage console
        ]
    )
    
    print("🚀 Démarrage du générateur EBIOS RM...")
    print("=" * 60)
    
    # Initialiser le générateur
    print("🔧 Initialisation du générateur...")
    generator = EBIOSTemplateGenerator()
    
    # Définir le chemin de sortie

    output_path = Path("c:/Users/mushm/Documents/AR/templates/ebios_risk_assessment_FR.xlsx")
    
    print(f"📁 Création du répertoire de sortie...")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    print(f"📊 Génération du template EBIOS RM...")
    print(f"   Destination: {output_path}")
    
    try:
        # Générer le template complet
        generator.generate_template(output_path)
        
        print("\n" + "=" * 60)
        print("✅ SUCCÈS : Template EBIOS RM généré avec succès!")
        print("=" * 60)
        print(f"📁 Fichier créé : {output_path}")
        print(f"📊 Taille du fichier : {output_path.stat().st_size / 1024:.1f} KB")
        
        # Vérifier la structure créée
        try:
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            sheet_names = wb.sheetnames
            print(f"📋 Onglets créés ({len(sheet_names)}) :")
            for i, sheet in enumerate(sheet_names, 1):
                print(f"   {i}. {sheet}")
            wb.close()
        except Exception as e:
            print(f"⚠️  Impossible de vérifier la structure : {e}")
        
        print("\n🎯 Le template est prêt pour utilisation!")
        print("   Vous pouvez maintenant exécuter 'python visualize_template.py'")
        
    except Exception as e:
        print("\n" + "=" * 60)
        print("❌ ERREUR lors de la génération du template")
        print("=" * 60)
        print(f"💥 Erreur : {e}")
        logging.exception("Erreur détaillée")
        print("\n💡 Suggestions de résolution :")
        print("   • Vérifiez que vous avez les droits d'écriture")
        print("   • Fermez Excel s'il est ouvert")
        print("   • Vérifiez l'espace disque disponible")
        return False    
    return True

if __name__ == "__main__":
    success = main()
    if not success:
        exit(1)