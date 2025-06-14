"""
Script de test minimal pour diagnostiquer les problèmes de génération.
"""

from pathlib import Path
import logging

def test_basic_generation():
    """Test basique de génération du template."""
    print("🧪 Test basique de génération...")
    
    try:
        from openpyxl import Workbook
        print("✅ OpenPyXL importé avec succès")
        
        # Test création simple
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        ws["A1"] = "Test EBIOS RM"
        
        # Test sauvegarde
        test_path = Path("c:/Users/mushm/Documents/AR/templates/test_simple.xlsx")
        test_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb.save(test_path)
        wb.close()
        
        print(f"✅ Fichier test créé : {test_path}")
        print(f"📊 Taille : {test_path.stat().st_size} bytes")
        
        # Nettoyer
        test_path.unlink()
        print("✅ Test basique réussi")
        
        return True
        
    except Exception as e:
        print(f"❌ Erreur test basique : {e}")
        return False

def test_import_generator():
    """Test d'import du générateur principal."""
    print("🧪 Test d'import du générateur...")
    
    try:
        from generate_template import EBIOSTemplateGenerator
        print("✅ EBIOSTemplateGenerator importé")
        
        generator = EBIOSTemplateGenerator()
        print("✅ Générateur instancié")
        
        print(f"📊 Nombre de tables de référence : {len(generator.reference_data)}")
        for table_name in generator.reference_data.keys():
            print(f"   • {table_name}")
        
        return True
        
    except Exception as e:
        print(f"❌ Erreur import générateur : {e}")
        import traceback
        traceback.print_exc()
        return False

def test_full_generation():
    """Test de génération complète."""
    print("🧪 Test de génération complète...")
    
    try:
        from generate_template import EBIOSTemplateGenerator
        
        generator = EBIOSTemplateGenerator()
        output_path = Path("c:/Users/mushm/Documents/AR/templates/test_full.xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        print("🔄 Génération en cours...")
        generator.generate_template(output_path)
        
        print(f"✅ Génération réussie : {output_path}")
        print(f"📊 Taille : {output_path.stat().st_size / 1024:.1f} KB")
        
        return True
        
    except Exception as e:
        print(f"❌ Erreur génération complète : {e}")
        import traceback
        traceback.print_exc()
        return False

def test_no_removed_records():
    """Test pour s'assurer qu'aucune formule ne sera supprimée par Excel."""
    print("🧪 Test de détection des formules qui causent 'Removed Records'...")
    
    try:
        from openpyxl import load_workbook
        
        test_path = Path("c:/Users/mushm/Documents/AR/templates/ebios_risk_assessment_FR.xlsx")
        
        if not test_path.exists():
            print("❌ Template non trouvé pour le test")
            return False
        
        print("🔍 Chargement et analyse du template...")
        wb = load_workbook(test_path, data_only=False)
        
        # Tables qui doivent exister pour les formules
        required_tables = ["Incidents"]
        missing_tables = []
        
        for table_name in required_tables:
            if table_name not in wb.sheetnames:
                missing_tables.append(table_name)
        
        if missing_tables:
            print(f"❌ Tables manquantes: {missing_tables}")
            return False
        
        # Vérifier les formules dangereuses
        dangerous_formulas = []
        critical_sheets = ["Synthese", "Dashboard_KPI", "Tendances_Evolutives"]
        
        for sheet_name in critical_sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            # Rechercher les références à des tables inexistantes
                            if any(ref in cell.value for ref in ["Personnel[", "Maturite["]):
                                dangerous_formulas.append({
                                    "sheet": sheet_name,
                                    "cell": cell.coordinate,
                                    "formula": cell.value[:50] + "..."
                                })
        
        if dangerous_formulas:
            print(f"⚠️ {len(dangerous_formulas)} formule(s) potentiellement dangereuse(s):")
            for formula in dangerous_formulas[:3]:  # Afficher les 3 premières
                print(f"   {formula['sheet']}.{formula['cell']}: {formula['formula']}")
            print("💡 Ces formules pourraient être supprimées par Excel lors de l'ouverture")
            return False
        
        print("✅ Aucune formule dangereuse détectée")
        wb.close()
        return True
        
    except Exception as e:
        print(f"❌ Erreur lors du test : {e}")
        return False

def main():
    """Exécute tous les tests de diagnostic."""
    print("🔍 DIAGNOSTIC EBIOS RM GENERATOR")
    print("=" * 50)
    
    tests = [
        ("Test basique OpenPyXL", test_basic_generation),
        ("Test import générateur", test_import_generator),
        ("Test génération complète", test_full_generation),
        ("Test formules sans 'Removed Records'", test_no_removed_records),
    ]
    
    results = []
    for test_name, test_func in tests:
        print(f"\n{test_name}:")
        print("-" * 30)
        
        try:
            result = test_func()
            results.append((test_name, result))
        except Exception as e:
            print(f"❌ Erreur inattendue : {e}")
            results.append((test_name, False))
    
    # Résumé
    print("\n" + "=" * 50)
    print("📋 RÉSUMÉ DES TESTS")
    print("=" * 50)
    
    for test_name, success in results:
        status = "✅ SUCCÈS" if success else "❌ ÉCHEC"
        print(f"{status} - {test_name}")
    
    total_success = sum(1 for _, success in results if success)
    print(f"\n🎯 Tests réussis : {total_success}/{len(results)}")
    
    if total_success == len(results):
        print("🎉 Tous les tests sont réussis!")
        print("💡 Le template ne devrait plus afficher de message 'Removed Records'")
    else:
        print("⚠️ Certains tests ont échoué")
        print("💡 Vérifiez les erreurs ci-dessus")

if __name__ == "__main__":
    main()
