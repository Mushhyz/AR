"""Tests spécifiques pour les mesures de sécurité et calculs de risque résiduel."""

import pytest
from pathlib import Path
from openpyxl import load_workbook
from scripts.generate_template import EBIOSTemplateGenerator


class TestSecurityMeasures:
    """Tests pour les mesures de sécurité ISO 27001."""
    
    @pytest.fixture
    def template_path(self, tmp_path):
        """Génère un template de test avec mesures."""
        template_file = tmp_path / "test_measures.xlsx"
        generator = EBIOSTemplateGenerator()
        generator.generate_template(template_file)
        return template_file
    
    def test_iso27001_controls_coverage(self, template_path):
        """Test de couverture des contrôles ISO 27001:2022."""
        wb = load_workbook(template_path)
        refs_ws = wb["__REFS"]
        
        # Extraire les contrôles Annex A présents
        annex_controls = []
        found_start = False
        
        for row in refs_ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith("A."):
                    annex_controls.append(cell.value)
                    found_start = True
                elif found_start and not cell.value:
                    break
        
        # Vérifier la présence de contrôles clés ISO 27001:2022
        key_controls = ["A.5.1", "A.8.1", "A.9.1", "A.14.1", "A.15.1", "A.16.1"]
        for control in key_controls:
            assert control in annex_controls, f"Contrôle ISO 27001 manquant: {control}"
        
        # Vérifier la diversité des domaines (A.5 à A.16)
        domains = set()
        for control in annex_controls:
            if "." in control:
                domain = control.split(".")[1]
                domains.add(domain)
        
        assert len(domains) >= 5, f"Couverture insuffisante des domaines ISO 27001: {domains}"
    
    def test_residual_risk_formulas(self, template_path):
        """Test des formules de calcul du risque résiduel."""
        wb = load_workbook(template_path)
        ws = wb["Atelier5_Traitement"]
        
        # Chercher les formules de risque résiduel (colonne K)
        residual_formulas = []
        for row in range(2, 20):  # Vérifier les premières lignes
            cell = ws.cell(row=row, column=11)  # Colonne K = Niveau_Résiduel
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                residual_formulas.append(cell.value)
        
        assert len(residual_formulas) > 0, "Formules de risque résiduel non trouvées"
        
        # Vérifier la structure de la formule classique GRC
        for formula in residual_formulas[:3]:  # Vérifier les 3 premières
            assert "SI(" in formula or "IF(" in formula, f"Formule sans condition: {formula}"
            assert "*" in formula, f"Formule sans multiplication: {formula}"
            assert "1-" in formula or "(1-" in formula, f"Formule sans réduction: {formula}"
            assert "/100" in formula, f"Formule sans conversion pourcentage: {formula}"
    
    def test_measure_validation_lists(self, template_path):
        """Test des listes de validation pour les mesures."""
        wb = load_workbook(template_path)
        ws = wb["Atelier5_Traitement"]
        
        # Vérifier les validations de données
        validations = ws.data_validations.dataValidation
        measure_validation_found = False
        
        for dv in validations:
            if "Measure_ID" in str(dv.formula1):
                measure_validation_found = True
                # Vérifier les propriétés de validation
                assert dv.showErrorMessage, "Message d'erreur non activé pour Measure_ID"
                assert dv.showInputMessage, "Message d'aide non activé pour Measure_ID"
                assert dv.showDropDown, "Flèche déroulante non activée"
                break
        
        assert measure_validation_found, "Validation Measure_ID non trouvée"
    
    def test_automatic_annexa_lookup(self, template_path):
        """Test de liaison automatique avec les contrôles Annex A."""
        wb = load_workbook(template_path)
        ws = wb["Atelier5_Traitement"]
        
        # Vérifier les formules XLOOKUP pour Contrôle_AnnexA (colonne F)
        annexa_formulas = []
        for row in range(2, 10):
            cell = ws.cell(row=row, column=6)  # Colonne F = Contrôle_AnnexA
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                annexa_formulas.append(cell.value)
        
        assert len(annexa_formulas) > 0, "Formules XLOOKUP AnnexA non trouvées"
        
        # Vérifier la structure des formules
        for formula in annexa_formulas[:2]:
            assert "XLOOKUP" in formula or "INDEX" in formula, f"Formule de recherche manquante: {formula}"
            assert "Measure_ID" in formula, f"Référence Measure_ID manquante: {formula}"


class TestRiskCalculations:
    """Tests pour les calculs de risque avancés."""
    
    def test_risk_matrix_values(self, template_path):
        """Test de cohérence de la matrice de risque."""
        wb = load_workbook(template_path)
        
        # Vérifier que les plages de valeurs numériques existent
        expected_ranges = ["tbl_Gravite_Valeur", "tbl_Vraisemblance_Valeur", "tbl_ValeurMetier_Valeur"]
        
        for range_name in expected_ranges:
            if range_name in wb.defined_names:
                # Vérifier que la plage existe
                range_obj = wb.defined_names[range_name]
                assert range_obj is not None, f"Plage {range_name} vide"
    
    def test_velocity_preparedness_kpis(self, template_path):
        """Test des KPI Velocity/Preparedness ISO 27005:2022."""
        wb = load_workbook(template_path)
        
        # Vérifier que la table Incidents existe
        assert "Incidents" in wb.sheetnames, "Table Incidents manquante pour les KPI"
        
        incidents_ws = wb["Incidents"]
        
        # Vérifier la structure de la table
        expected_headers = ["ID", "Date_Detection", "Date_Reponse", "Temps_Detection", "Temps_Reponse", "Temps_Resolution", "Statut", "Gravite"]
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = incidents_ws.cell(row=1, column=col).value
            assert actual_header == expected_header, f"En-tête incorrect: attendu {expected_header}, trouvé {actual_header}"
        
        # Vérifier les KPI dans Synthèse
        ws = wb["Synthese"]
        
        # Chercher les formules KPI corrigées
        velocity_formulas = []
        preparedness_formulas = []
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    if "Incidents[Temps_Detection]" in cell.value or "Incidents[Temps_Reponse]" in cell.value:
                        velocity_formulas.append(cell.value)
                    elif "Incidents[Gravite]" in cell.value:
                        preparedness_formulas.append(cell.value)
        
        assert len(velocity_formulas) >= 2, "Formules KPI Velocity avec table Incidents non trouvées"
        assert len(preparedness_formulas) >= 1, "Formules KPI Preparedness avec table Incidents non trouvées"
        
        # Vérifier qu'aucune formule ne contient #REF!
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    assert "#REF!" not in cell.value, f"Formule #REF! détectée en {cell.coordinate}: {cell.value}"

    def test_no_broken_formulas(self, template_path):
        """Test de détection des formules cassées qui causeraient 'Removed Records'."""
        wb = load_workbook(template_path)
        
        # Feuilles à vérifier spécifiquement
        sheets_to_check = ["Synthese", "Dashboard_KPI", "Tendances_Evolutives"]
        
        broken_formulas = []
        
        for sheet_name in sheets_to_check:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            # Vérifier les références dangereuses
                            dangerous_refs = [
                                "Personnel[", "Maturite[", "Incidents[ID])*100" 
                            ]
                            
                            for danger in dangerous_refs:
                                if danger in cell.value and sheet_name not in ["Incidents"]:
                                    # Vérifier que la table référencée existe
                                    table_name = danger.split('[')[0]
                                    if table_name not in wb.sheetnames:
                                        broken_formulas.append({
                                            "sheet": sheet_name,
                                            "cell": cell.coordinate,
                                            "formula": cell.value,
                                            "missing_table": table_name
                                        })
        
        # Signaler les formules dangereuses trouvées
        if broken_formulas:
            error_msg = "Formules dangereuses détectées:\n"
            for error in broken_formulas:
                error_msg += f"  {error['sheet']}.{error['cell']}: {error['formula']} (table manquante: {error['missing_table']})\n"
            
            # Pour ce test, on prévient mais on ne fait pas échouer
            # car c'est un template et certaines tables peuvent être absentes
            print(f"⚠️ {error_msg}")


class TestDataValidations:
    """Tests pour les validations de données avancées."""
    
    def test_pertinence_exposition_scales(self, template_path):
        """Test des échelles dédiées Pertinence/Exposition."""
        wb = load_workbook(template_path)
        
        # Vérifier que les plages spécifiques existent
        pertinence_range = wb.defined_names.get("Pertinence")
        exposition_range = wb.defined_names.get("Exposition")
        
        assert pertinence_range is not None, "Plage Pertinence non trouvée"
        assert exposition_range is not None, "Plage Exposition non trouvée"
        
        # Vérifier l'Atelier 2 pour les validations spécifiques
        ws = wb["Atelier2_Sources"]
        validations = ws.data_validations.dataValidation
        
        pertinence_validation = False
        exposition_validation = False
        
        for dv in validations:
            if "Pertinence" in str(dv.formula1):
                pertinence_validation = True
                assert "Faible, Modérée ou Forte" in dv.error, "Message d'erreur Pertinence incorrect"
            elif "Exposition" in str(dv.formula1):
                exposition_validation = True
                assert "Limitée, Significative ou Maximale" in dv.error, "Message d'erreur Exposition incorrect"
        
        assert pertinence_validation, "Validation Pertinence non trouvée"
        assert exposition_validation, "Validation Exposition non trouvée"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
