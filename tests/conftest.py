"""Configuration pytest pour les tests EBIOS RM."""

import pytest
import tempfile
import shutil
from pathlib import Path


@pytest.fixture(scope="session")
def test_config_dir():
    """Crée un répertoire de configuration temporaire pour les tests."""
    temp_dir = Path(tempfile.mkdtemp())

    # Créer des fichiers CSV de test
    assets_csv = temp_dir / "assets.csv"
    assets_csv.write_text(
        """id,type,label,criticality
    A001,Data,Customer Database,Critical
    A002,System,Web Server,High
    A003,Application,Mobile App,Medium
    """,
        encoding="utf-8",
    )

    threats_csv = temp_dir / "threats.csv"
    threats_csv.write_text(
        """sr_id,ov_id,strategic_path,operational_steps
    SR001,OV001,External Attack,Step1:High,Step2:Medium
    SR002,OV002,Internal Fraud,Step1:Low,Step2:High,Step3:Medium
    """,
        encoding="utf-8",
    )

    settings_yaml = temp_dir / "settings.yaml"
    settings_yaml.write_text(
        """excel_template: templates/ebiosrm_empty.xlsx
    severity_scale: [Low, Medium, High, Critical]
    likelihood_scale: [One-shot, Occasional, Probable, Systematic]
    output_dir: build/
    """,
        encoding="utf-8",
    )

    yield temp_dir

    # Cleanup
    shutil.rmtree(temp_dir)


@pytest.fixture
def clean_output_dir(tmp_path):
    """Crée un répertoire de sortie propre pour chaque test."""
    output_dir = tmp_path / "output"
    output_dir.mkdir(exist_ok=True)
    return output_dir


@pytest.fixture
def sample_template_path(tmp_path):
    """Crée un template Excel minimal pour les tests."""
    try:
        from scripts.generate_template import EBIOSTemplateGenerator

        template_file = tmp_path / "sample_template.xlsx"
        generator = EBIOSTemplateGenerator()
        generator.generate_template(template_file)
        return template_file
    except ImportError:
        pytest.skip("Generator non disponible pour ce test")


@pytest.fixture(autouse=True)
def configure_logging():
    """Configure le logging pour les tests."""
    import logging

    logging.basicConfig(level=logging.WARNING)  # Réduire le bruit pendant les tests


"""Test configuration and shared fixtures."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest
import yaml

from ebiosrm_core.models import Asset, Threat, Settings


@pytest.fixture
def sample_assets():
    """Sample assets for testing."""
    return [
        Asset(
            id="A001", type="Data", label="Customer Database", criticality="Critical"
        ),
        Asset(id="A002", type="System", label="Web Server", criticality="High"),
        Asset(id="A003", type="Data", label="Application Logs", criticality="Low"),
        Asset(
            id="A004", type="System", label="Database Server", criticality="Critical"
        ),
        Asset(
            id="A005", type="Network", label="Internal Network", criticality="Medium"
        ),
    ]


@pytest.fixture
def sample_threats():
    """Sample threats for testing."""
    return [
        Threat(
            sr_id="SR001",
            ov_id="OV001",
            strategic_path="External Cyber Attack",
            operational_steps="Reconnaissance:Low,Initial Access:Medium,Privilege Escalation:High,Data Exfiltration:Critical",
        ),
        Threat(
            sr_id="SR002",
            ov_id="OV002",
            strategic_path="Insider Threat",
            operational_steps="Credential Abuse:High,Data Access:Critical,Data Theft:Critical",
        ),
        Threat(
            sr_id="SR003",
            ov_id="OV003",
            strategic_path="Supply Chain Attack",
            operational_steps="Third Party Compromise:Medium,Lateral Movement:High,Persistence:High",
        ),
    ]


@pytest.fixture
def sample_objectives():
    """Sample objectives for testing - aligned with EBIOS RM targeted objectives."""
    from ebiosrm_core.models import TargetedObjective

    return [
        TargetedObjective(
            id="OBJ001",
            label="Vol de données clients",
            target_assets=["A001", "A003"],
            business_impact="Critical",
            attack_scenarios=["SR001", "SR002"],
        ),
        TargetedObjective(
            id="OBJ002",
            label="Indisponibilité des services",
            target_assets=["A002", "A008"],
            business_impact="High",
            attack_scenarios=["SR006"],
        ),
        TargetedObjective(
            id="OBJ003",
            label="Atteinte à la réputation",
            target_assets=["A001", "A005"],
            business_impact="Medium",
            attack_scenarios=["SR002", "SR005"],
        ),
    ]


@pytest.fixture
def sample_risk_sources():
    """Sample risk sources aligned with EBIOS RM methodology."""
    from ebiosrm_core.models import RiskSource

    return [
        RiskSource(
            id="RS001",
            label="Cybercriminels organisés",
            category="Criminalité organisée",
            motivation="Gain financier",
            capability_level="High",
            resources="Outils avancés, compétences techniques",
        ),
        RiskSource(
            id="RS002",
            label="Acteurs étatiques",
            category="Espionnage d'État",
            motivation="Intelligence économique",
            capability_level="Critical",
            resources="Ressources gouvernementales illimitées",
        ),
        RiskSource(
            id="RS003",
            label="Employés malveillants",
            category="Menace interne",
            motivation="Vengeance ou gain personnel",
            capability_level="Medium",
            resources="Accès privilégié interne",
        ),
    ]


@pytest.fixture
def sample_stakeholders():
    """Sample stakeholders for EBIOS RM process."""
    from ebiosrm_core.models import Stakeholder

    return [
        Stakeholder(
            id="ST001",
            name="Direction Générale",
            type="Internal",
            role="Pilotage stratégique",
            responsibilities=[
                "Validation des orientations",
                "Allocation des ressources",
            ],
            contact_info="direction@entreprise.fr",
        ),
        Stakeholder(
            id="ST002",
            name="RSSI",
            type="Internal",
            role="Responsable sécurité",
            responsibilities=["Animation EBIOS RM", "Définition politique sécurité"],
            contact_info="rssi@entreprise.fr",
        ),
        Stakeholder(
            id="ST003",
            name="ANSSI",
            type="Regulatory",
            role="Autorité de tutelle",
            responsibilities=["Réglementation", "Contrôle conformité"],
            contact_info="contact@ssi.gouv.fr",
        ),
    ]


@pytest.fixture
def ebios_workshop_data():
    """Complete EBIOS RM workshop data structure for testing."""
    return {
        "atelier_1": {
            "mission_metier": "Fourniture de services numériques aux clients",
            "actifs_supports": ["Serveurs", "Réseaux", "Applications"],
            "actifs_essentiels": ["Données clients", "Services en ligne"],
            "parties_prenantes": ["Clients", "Fournisseurs", "Régulateurs"],
            "sources_menaces": ["Cybercriminels", "Concurrents", "États"],
        },
        "atelier_2": {
            "ecosysteme": ["Partenaires", "Sous-traitants", "Cloud providers"],
            "cartographie_sr": {
                "RS001": {"motivation": "Financière", "capacites": "Élevées"},
                "RS002": {"motivation": "Espionnage", "capacites": "Critiques"},
            },
        },
        "atelier_3": {
            "scenarios_strategiques": {
                "SR001": {
                    "source": "RS001",
                    "objectifs": ["OBJ001", "OBJ002"],
                    "gravite": "Critical",
                    "vraisemblance": "Probable",
                }
            }
        },
        "atelier_4": {
            "scenarios_operationnels": {
                "OV001": {
                    "scenario_strategique": "SR001",
                    "vecteurs_attaque": ["Email", "Web", "USB"],
                    "etapes": [
                        "Reconnaissance",
                        "Intrusion",
                        "Persistance",
                        "Exfiltration",
                    ],
                }
            }
        },
        "atelier_5": {
            "mesures_securite": {
                "existantes": ["Antivirus", "Firewall", "Sauvegarde"],
                "residuelles": ["Formation", "Chiffrement", "Monitoring"],
                "options_traitement": ["Réduire", "Éviter", "Transférer", "Accepter"],
            }
        },
    }


@pytest.fixture
def pme_profile_data():
    """Simplified data profile for PME/TPE organizations."""
    return {
        "criteres_impact": ["Négligeable", "Limité", "Important", "Critique"],
        "niveaux_vraisemblance": ["Minimal", "Significatif", "Élevé", "Maximal"],
        "types_actifs": ["Données", "Systèmes", "Locaux", "Personnel"],
        "categories_menaces": ["Cybercriminalité", "Espionnage", "Sabotage", "Erreur"],
        "mesures_types": [
            "Organisationnelles",
            "Techniques",
            "Physiques",
            "Juridiques",
        ],
        "referentiels": ["ISO 27001", "NIST", "ANSSI", "Secteur spécifique"],
    }


@pytest.fixture
def sample_settings():
    """Sample settings for testing."""
    return Settings(
        excel_template="ebiosrm_template.xlsx",
        output_dir="output/",
        severity_scale=["Low", "Medium", "High", "Critical"],
        likelihood_scale=["One-shot", "Occasional", "Probable", "Systematic"],
    )


@pytest.fixture
def temp_config_dir(
    tmp_path, sample_assets, sample_threats, sample_objectives, sample_settings
):
    """Create temporary configuration directory with test data."""
    config_dir = tmp_path / "config"
    config_dir.mkdir()

    # Create assets.csv with proper enum values (not string representation)
    assets_file = config_dir / "assets.csv"
    with open(assets_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "type", "label", "criticality"])
        for asset in sample_assets:
            writer.writerow(
                [asset.id, asset.type, asset.label, asset.criticality.value]
            )

    # Create threats.csv with proper string formatting
    threats_file = config_dir / "threats.csv"
    with open(threats_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["sr_id", "ov_id", "strategic_path", "operational_steps"])
        for threat in sample_threats:
            writer.writerow(
                [
                    threat.sr_id,
                    threat.ov_id,
                    threat.strategic_path,
                    threat.operational_steps,
                ]
            )

    # Create objectives.csv with the working structure
    objectives_file = config_dir / "objectives.csv"
    with open(objectives_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            ["id", "label", "target_assets", "business_impact", "attack_scenarios"]
        )
        for objective in sample_objectives:
            # Convert lists to comma-separated strings
            target_assets_str = (
                ",".join(objective.target_assets)
                if isinstance(objective.target_assets, list)
                else str(objective.target_assets)
            )
            scenarios_str = (
                ",".join(objective.attack_scenarios)
                if isinstance(objective.attack_scenarios, list)
                else str(objective.attack_scenarios)
            )
            writer.writerow(
                [
                    objective.id,
                    objective.label,
                    target_assets_str,
                    objective.business_impact.value,
                    scenarios_str,
                ]
            )

    # Create minimal optional files to prevent FileNotFoundError
    for optional_file in ["risk_sources.csv", "stakeholders.csv", "measures.csv"]:
        file_path = config_dir / optional_file
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if optional_file == "risk_sources.csv":
                writer.writerow(
                    [
                        "id",
                        "label",
                        "category",
                        "motivation",
                        "capability_level",
                        "resources",
                    ]
                )
            elif optional_file == "stakeholders.csv":
                writer.writerow(
                    ["id", "name", "type", "role", "responsibilities", "contact_info"]
                )
            elif optional_file == "measures.csv":
                writer.writerow(
                    [
                        "id",
                        "label",
                        "type",
                        "description",
                        "effectiveness",
                        "implementation_cost",
                        "target_threats",
                        "responsible_stakeholder",
                    ]
                )

    # Create settings.yaml using model_dump instead of dict()
    settings_file = config_dir / "settings.yaml"
    with open(settings_file, "w", encoding="utf-8") as f:
        yaml.dump(sample_settings.model_dump(), f)

    return config_dir


@pytest.fixture
def cli_runner():
    """Provide a CLI test runner."""
    from typer.testing import CliRunner

    return CliRunner()


@pytest.fixture
def temp_output_dir(tmp_path):
    """Create temporary output directory for testing."""
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    return output_dir


@pytest.fixture
def sample_excel_template(tmp_path):
    """Create a minimal Excel template for testing."""
    from openpyxl import Workbook

    template_path = tmp_path / "ebiosrm_template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "EBIOS_RM"

    # Add some basic headers
    ws["A1"] = "Asset ID"
    ws["B1"] = "Asset Type"
    ws["C1"] = "Asset Label"
    ws["D1"] = "Criticality"

    wb.save(template_path)
    return template_path


@pytest.fixture
def complete_test_environment(temp_config_dir, temp_output_dir, sample_excel_template):
    """Provide a complete test environment with all necessary files."""
    return {
        "config_dir": temp_config_dir,
        "output_dir": temp_output_dir,
        "template": sample_excel_template,
    }


@pytest.fixture
def mock_cli_app():
    """Provide a mock CLI app for testing when the real CLI isn't available."""
    import typer

    app = typer.Typer()

    @app.command()
    def validate():
        """Mock validate command."""
        typer.echo("Validation successful")
        return True

    @app.command()
    def export(fmt: str = "json"):
        """Mock export command."""
        typer.echo(f"Export to {fmt} format successful")
        return True

    return app


@pytest.fixture
def project_root():
    """Get the project root directory."""
    return Path(__file__).parent.parent


@pytest.fixture
def debug_module_info(project_root):
    """Debug fixture to help identify the correct module structure."""
    import sys
    import importlib.util

    info = {
        "project_root": str(project_root),
        "python_path": sys.path,
        "available_modules": [],
    }

    # Check for ebiosrm modules
    for module_name in ["ebiosrm", "ebiosrm_core", "ebiosrm_generator"]:
        try:
            spec = importlib.util.find_spec(module_name)
            if spec:
                info["available_modules"].append(
                    {
                        "name": module_name,
                        "location": spec.origin,
                        "submodule_search_paths": spec.submodule_search_paths,
                    }
                )
        except (ImportError, ModuleNotFoundError):
            pass

    return info


@pytest.fixture
def debug_csv_data(temp_config_dir):
    """Debug fixture to inspect CSV data structure."""
    import csv

    threats_file = temp_config_dir / "threats.csv"
    debug_info = {
        "file_exists": threats_file.exists(),
        "file_content": [],
        "csv_rows": [],
    }

    if threats_file.exists():
        # Read raw file content
        with open(threats_file, "r", encoding="utf-8") as f:
            debug_info["file_content"] = f.readlines()

        # Read CSV structure
        with open(threats_file, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader):
                debug_info["csv_rows"].append(
                    {
                        "row_number": i,
                        "keys": list(row.keys()),
                        "key_types": {k: type(k).__name__ for k in row.keys()},
                        "values": dict(row),
                        "value_types": {k: type(v).__name__ for k, v in row.items()},
                    }
                )

    return debug_info


@pytest.fixture
def csv_debugging_helper():
    """Helper to debug CSV loading issues in real config files."""

    def debug_csv(file_path):
        import csv

        print(f"Debugging {file_path}")

        with open(file_path, "r", encoding="utf-8") as f:
            # Read raw content
            content = f.read()
            print(f"Raw content (first 200 chars): {repr(content[:200])}")

        with open(file_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            print(f"Headers: {reader.fieldnames}")
            print(f"Header types: {[type(h) for h in reader.fieldnames]}")

            for i, row in enumerate(reader):
                if i < 3:  # Show first 3 rows
                    print(f"Row {i}: {dict(row)}")
                    print(f"Key types: {[(k, type(k)) for k in row.keys()]}")
                break  # Only show first row for debugging

    return debug_csv


@pytest.fixture
def real_config_debugger():
    """Debug fixture specifically for the real config directory."""

    def debug_real_config():
        from pathlib import Path
        import csv

        config_path = Path("config")
        threats_file = config_path / "threats.csv"

        if not threats_file.exists():
            print(f"❌ {threats_file} does not exist")
            return

        print(f"🔍 Debugging {threats_file}")

        # Check raw file content
        with open(threats_file, "rb") as f:
            raw_bytes = f.read(100)
            print(f"Raw bytes: {raw_bytes}")

        # Check text content
        with open(threats_file, "r", encoding="utf-8") as f:
            lines = f.readlines()[:5]
            print(f"First 5 lines: {lines}")

        # Check CSV parsing
        try:
            with open(threats_file, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                print(f"Headers: {reader.fieldnames}")
                print(f"Header types: {[(h, type(h)) for h in reader.fieldnames]}")

                for i, row in enumerate(reader):
                    print(f"Row {i} keys: {[(k, type(k)) for k in row.keys()]}")
                    print(f"Row {i} values: {dict(row)}")
                    if i >= 2:  # Only show first 3 rows
                        break
        except Exception as e:
            print(f"❌ CSV parsing error: {e}")

    return debug_real_config


@pytest.fixture
def csv_header_cleaner():
    """Fixture to clean CSV headers and ensure they are valid strings."""

    def clean_csv_file(file_path):
        import csv
        import tempfile

        # Read original file with different encodings to handle BOM
        encodings_to_try = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
        content = None

        for encoding in encodings_to_try:
            try:
                with open(file_path, "r", encoding=encoding) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue

        if content is None:
            raise ValueError(
                f"Could not decode {file_path} with any supported encoding"
            )

        # Remove BOM if present
        if content.startswith("\ufeff"):
            content = content[1:]

        # Write to temp file and read back
        with tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv", encoding="utf-8"
        ) as temp_f:
            temp_f.write(content)
            temp_path = temp_f.name

        try:
            # Read and validate CSV structure
            with open(temp_path, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames

                # Check for problematic headers
                if headers is None:
                    raise ValueError("No headers found in CSV")

                clean_headers = []
                for i, h in enumerate(headers):
                    if h is None:
                        clean_headers.append(f"unnamed_column_{i}")
                    elif not isinstance(h, str):
                        clean_headers.append(str(h).strip())
                    else:
                        clean_headers.append(h.strip())

                # Remove empty headers
                clean_headers = [h for h in clean_headers if h]

                # Read all rows
                rows = []
                for row in reader:
                    clean_row = {}
                    for old_h, new_h in zip(headers, clean_headers):
                        if old_h is None:
                            continue
                        value = row.get(old_h, "")
                        clean_row[new_h] = (
                            str(value).strip() if value is not None else ""
                        )
                    rows.append(clean_row)

            # Write cleaned version back to original file
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                if rows and clean_headers:
                    writer = csv.DictWriter(f, fieldnames=clean_headers)
                    writer.writeheader()
                    writer.writerows(rows)

            return clean_headers

        finally:
            # Clean up temp file
            import os

            try:
                os.unlink(temp_path)
            except FileNotFoundError:
                pass

    return clean_csv_file


@pytest.fixture
def validate_real_config():
    """Validate and fix the real config directory CSV files."""

    def validate_config_dir():
        from pathlib import Path

        config_path = Path("config")

        if not config_path.exists():
            print("❌ Config directory does not exist")
            return False

        csv_files = ["threats.csv", "assets.csv"]
        issues = []

        for csv_file in csv_files:
            file_path = config_path / csv_file
            if file_path.exists():
                try:
                    # Try to validate the CSV
                    import csv

                    with open(file_path, "r", newline="", encoding="utf-8") as f:
                        reader = csv.DictReader(f)
                        headers = reader.fieldnames

                        if headers is None:
                            issues.append(f"{csv_file}: No headers found")
                            continue

                        # Check for problematic headers
                        for i, header in enumerate(headers):
                            if header is None:
                                issues.append(
                                    f"{csv_file}: None header at position {i}"
                                )
                            elif not isinstance(header, str):
                                issues.append(
                                    f"{csv_file}: Non-string header '{header}' (type: {type(header)})"
                                )

                        # Try to read first row
                        try:
                            first_row = next(reader, None)
                            if first_row:
                                for key in first_row.keys():
                                    if not isinstance(key, str):
                                        issues.append(
                                            f"{csv_file}: Non-string key '{key}' (type: {type(key)})"
                                        )
                        except Exception as e:
                            issues.append(f"{csv_file}: Error reading first row: {e}")

                except Exception as e:
                    issues.append(f"{csv_file}: Error opening file: {e}")

        if issues:
            print("❌ CSV validation issues found:")
            for issue in issues:
                print(f"  - {issue}")
            return False
        else:
            print("✅ CSV files validated successfully")
            return True

    return validate_config_dir


@pytest.fixture
def csv_file_cleaner():
    """Fixture to clean problematic CSV files."""

    def clean_csv_file(file_path):
        """Clean a CSV file to remove common issues."""
        import csv

        # Read with BOM handling
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            content = f.read()

        # Remove any remaining BOM characters
        content = content.replace("\ufeff", "")

        # Write to temporary file
        temp_file = file_path.with_suffix(".cleaned.csv")
        with open(temp_file, "w", encoding="utf-8", newline="") as f:
            f.write(content)

        # Validate the cleaned file
        with open(temp_file, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames

            # Check headers
            if not headers or None in headers:
                raise ValueError("Invalid or missing headers in CSV")

            # Ensure all headers are strings
            clean_headers = [str(h).strip() for h in headers if h is not None]

            # Re-read and clean data
            f.seek(0)
            reader = csv.DictReader(f)
            rows = []
            for row in reader:
                clean_row = {}
                for old_key, new_key in zip(headers, clean_headers):
                    value = row.get(old_key, "")
                    clean_row[new_key] = str(value).strip() if value is not None else ""
                rows.append(clean_row)

        # Write final cleaned version
        with open(temp_file, "w", newline="", encoding="utf-8") as f:
            if rows:
                writer = csv.DictWriter(f, fieldnames=clean_headers)
                writer.writeheader()
                writer.writerows(rows)

        return temp_file

    return clean_csv_file


@pytest.fixture
def test_csv_with_embedded_commas(tmp_path):
    """Test fixture for CSV files with embedded commas in fields."""
    csv_content = '''sr_id,ov_id,strategic_path,operational_steps
SR001,OV001,External Attack,"Step1:High,Step2:Medium,Step3:Low"
SR002,OV002,Internal Threat,"Step1:Critical,Step2:High"'''

    csv_file = tmp_path / "test_threats.csv"
    with open(csv_file, "w", encoding="utf-8") as f:
        f.write(csv_content)

    return csv_file


@pytest.fixture
def working_config_validator():
    """Fixture to validate that config matches working structure."""

    def validate_working_config(config_dir):
        from pathlib import Path
        import csv

        validation_results = {
            "assets": {"exists": False, "valid": False, "count": 0},
            "threats": {"exists": False, "valid": False, "count": 0},
            "objectives": {"exists": False, "valid": False, "count": 0},
        }

        # Check assets.csv
        assets_file = Path(config_dir) / "assets.csv"
        if assets_file.exists():
            validation_results["assets"]["exists"] = True
            try:
                with open(assets_file, "r", newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    expected_headers = ["id", "type", "label", "criticality"]
                    if reader.fieldnames == expected_headers:
                        validation_results["assets"]["valid"] = True
                        validation_results["assets"]["count"] = sum(1 for _ in reader)
            except Exception:
                pass

        # Check threats.csv
        threats_file = Path(config_dir) / "threats.csv"
        if threats_file.exists():
            validation_results["threats"]["exists"] = True
            try:
                with open(threats_file, "r", newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    expected_headers = [
                        "sr_id",
                        "ov_id",
                        "strategic_path",
                        "operational_steps",
                    ]
                    if reader.fieldnames == expected_headers:
                        validation_results["threats"]["valid"] = True
                        validation_results["threats"]["count"] = sum(1 for _ in reader)
            except Exception:
                pass

        # Check objectives.csv
        objectives_file = Path(config_dir) / "objectives.csv"
        if objectives_file.exists():
            validation_results["objectives"]["exists"] = True
            try:
                with open(objectives_file, "r", newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    expected_headers = [
                        "id",
                        "label",
                        "target_assets",
                        "business_impact",
                        "attack_scenarios",
                    ]
                    if reader.fieldnames == expected_headers:
                        validation_results["objectives"]["valid"] = True
                        validation_results["objectives"]["count"] = sum(
                            1 for _ in reader
                        )
            except Exception:
                pass

        return validation_results

    return validate_working_config


@pytest.fixture
def excel_template_validator():
    """Validator for Excel template structure and formatting."""

    def validate_excel_template(workbook_path):
        from openpyxl import load_workbook

        validation_results = {
            "structure": {"valid": True, "issues": []},
            "data_validation": {"valid": True, "issues": []},
            "formatting": {"valid": True, "issues": []},
            "hidden_sheets": {"valid": True, "issues": []},
        }

        try:
            wb = load_workbook(workbook_path)

            # **CORRECTION** : Noms d'onglets conformes au générateur
            expected_sheets = [
                "Config_EBIOS",
                "Atelier1_Socle", 
                "Atelier2_Sources", 
                "Atelier3_Scenarios",
                "Atelier4_Operationnels", 
                "Atelier5_Traitement", 
                "Synthese"
            ]

            missing_sheets = [
                sheet for sheet in expected_sheets if sheet not in wb.sheetnames
            ]
            if missing_sheets:
                validation_results["structure"]["valid"] = False
                validation_results["structure"]["issues"].append(
                    f"Missing sheets: {missing_sheets}"
                )

            # **CORRECTION** : Vérifier l'onglet de références caché
            if "__REFS" not in wb.sheetnames:
                validation_results["hidden_sheets"]["valid"] = False
                validation_results["hidden_sheets"]["issues"].append(
                    "Reference sheet __REFS not found"
                )
            elif wb["__REFS"].sheet_state != "veryHidden":
                validation_results["hidden_sheets"]["valid"] = False
                validation_results["hidden_sheets"]["issues"].append(
                    "Reference sheet __REFS is not hidden"
                )

            # **CORRECTION** : Vérifier les validations sur les feuilles clés
            for sheet_name in ["Atelier3_Scenarios", "Atelier4_Operationnels"]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if (
                        not hasattr(ws, "data_validations")
                        or len(ws.data_validations.dataValidation) == 0
                    ):
                        validation_results["data_validation"]["valid"] = False
                        validation_results["data_validation"]["issues"].append(
                            f"No data validation found in {sheet_name}"
                        )
                    else:
                        # Vérifier que les validations ont showDropDown=False
                        dropdown_validation_found = False
                        for dv in ws.data_validations.dataValidation:
                            if dv.showDropDown == False:
                                dropdown_validation_found = True
                                break
                        
                        if not dropdown_validation_found:
                            validation_results["data_validation"]["valid"] = False
                            validation_results["data_validation"]["issues"].append(
                                f"No dropdown validation with showDropDown=False in {sheet_name}"
                            )

        except Exception as e:
            validation_results["structure"]["valid"] = False
            validation_results["structure"]["issues"].append(
                f"Error loading workbook: {str(e)}"
            )

        return validation_results

    return validate_excel_template


@pytest.fixture
def pme_profile_validator():
    """Fixture to validate PME (Small/Medium Enterprise) specific requirements."""

    def validate_pme_profile(risk_results):
        """Validate that risk results are suitable for PME context."""
        pme_issues = []

        # Check for simplified risk levels appropriate for PMEs
        risk_levels = [result["risk_level"] for result in risk_results]
        if "Critical" in risk_levels:
            critical_count = risk_levels.count("Critical")
            if critical_count > 3:
                pme_issues.append(
                    f"Too many critical risks ({critical_count}) for PME context"
                )

        # Check operational steps complexity
        for result in risk_results:
            steps = result["operational_steps"]
            if len(steps.split(",")) > 5:
                pme_issues.append(
                    f"Threat {result['threat_id']} has too many operational steps for PME"
                )

        # Validate measure effectiveness for PME budget
        return {
            "is_pme_suitable": len(pme_issues) == 0,
            "issues": pme_issues,
            "recommendation": "Simplify complexity for SME context"
            if pme_issues
            else "Suitable for PME",
        }

    return validate_pme_profile


@pytest.fixture
def risk_matrix_validator():
    """Fixture to validate risk matrix calculations."""

    def validate_risk_matrix(risk_results):
        """Validate that risk levels are correctly calculated."""
        validation_results = []

        for result in risk_results:
            threat_id = result["threat_id"]
            likelihood = result["likelihood_score"]
            severity = result["severity_score"]
            calculated_risk = result["risk_level"]

            # Manual risk matrix calculation for validation
            # Based on the logic in models.py
            matrix = [
                ["Low", "Low", "Medium", "High"],
                ["Low", "Medium", "Medium", "High"],
                ["Medium", "Medium", "High", "Critical"],
                ["Medium", "High", "Critical", "Critical"],
            ]

            sev_idx = min(int(severity) - 1, 3)
            lik_idx = min(int(likelihood) - 1, 3)
            expected_risk = matrix[sev_idx][lik_idx]

            validation_results.append(
                {
                    "threat_id": threat_id,
                    "likelihood": likelihood,
                    "severity": severity,
                    "calculated_risk": calculated_risk,
                    "expected_risk": expected_risk,
                    "is_correct": calculated_risk == expected_risk,
                }
            )

        return validation_results

    return validate_risk_matrix


@pytest.fixture
def operational_steps_fixer():
    """Fixture to fix truncated operational steps in threats.csv."""

    def fix_operational_steps():
        """Fix the operational steps that appear to be truncated."""

        # Updated operational steps with complete sequences
        fixed_steps = {
            "SR001": "Reconnaissance:Low,Initial Access:Medium,Persistence:High,Data Exfiltration:High",
            "SR002": "Access Granted:High,Data Collection:Medium,Data Theft:High",
            "SR003": "Building Access:Low,System Access:Medium,Data Access:High",
            "SR004": "Vendor Compromise:Medium,Software Distribution:High,System Infection:Critical",
            "SR005": "Information Gathering:Medium,Contact Target:High,Credential Extraction:Critical",
            "SR006": "Initial Infection:Medium,Lateral Movement:High,Encryption:Critical",
        }

        return fixed_steps

    return fix_operational_steps
