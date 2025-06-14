"""Integration tests for complete EBIOS RM workflow."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from pydantic import ValidationError

from ebiosrm_core import generator, loader
from ebiosrm_core.models import Asset, Threat


class TestDataLoading:
    """Test data loading functionality."""

    def test_load_all_success(self, temp_config_dir):
        """Test successful loading of all configuration data."""
        assets, threats, settings, risk_sources, objectives, stakeholders, measures = (
            loader.load_all(temp_config_dir)
        )

        assert len(assets) == 5  # Updated to match sample_assets fixture
        assert len(threats) == 3  # Updated to match sample_threats fixture
        assert settings.output_dir == "output/"  # Updated to match sample_settings

        # Verify first asset
        assert assets[0].id == "A001"
        assert assets[0].criticality.value == "Critical"

        # Verify first threat
        assert threats[0].sr_id == "SR001"
        assert "Reconnaissance:Low" in threats[0].operational_steps

    def test_load_missing_directory(self):
        """Test loading from non-existent directory."""
        with pytest.raises(FileNotFoundError):
            loader.load_all(Path("non_existent"))

    def test_load_invalid_csv_format(self, tmp_path):
        """Test loading CSV with invalid format."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()

        # Create invalid assets CSV (missing required columns)
        assets_file = config_dir / "assets.csv"
        with open(assets_file, "w") as f:
            f.write("id,wrong_column\nA001,value\n")

        # Create empty threats file to avoid FileNotFoundError
        threats_file = config_dir / "threats.csv"
        with open(threats_file, "w") as f:
            f.write("sr_id,ov_id,strategic_path,operational_steps\n")

        with pytest.raises(ValueError, match="Missing columns"):
            loader.load_all(config_dir)


class TestModels:
    """Test Pydantic models and business logic."""

    def test_asset_severity_score(self):
        """Test asset severity scoring."""
        asset_low = Asset(id="A1", type="Data", label="Test", criticality="Low")
        asset_critical = Asset(
            id="A2", type="System", label="Test", criticality="Critical"
        )

        assert asset_low.severity_score() == 1
        assert asset_critical.severity_score() == 4

    def test_threat_likelihood_calculation(self):
        """Test threat likelihood scoring."""
        threat = Threat(
            sr_id="SR001",
            ov_id="OV001",
            strategic_path="Test Attack",
            operational_steps="Step1:Low,Step2:High,Step3:Medium",
        )

        # Should be (1 + 3 + 2) / 3 = 2.0
        assert threat.likelihood_score() == 2.0

    def test_threat_risk_level_calculation(self):
        """Test risk level matrix calculation."""
        threat = Threat(
            sr_id="SR001",
            ov_id="OV001",
            strategic_path="Test",
            operational_steps="Step1:High,Step2:High",  # likelihood = 3.0
        )

        # Critical severity (4) + High likelihood (3) should give Critical risk
        risk_level = threat.risk_level(max_asset_severity=4)
        assert risk_level == "Critical"

    def test_invalid_threat_steps_format(self):
        """Test validation of operational steps format."""
        with pytest.raises(ValidationError, match="operational_steps must contain"):
            Threat(
                sr_id="SR001",
                ov_id="OV001",
                strategic_path="Test",
                operational_steps="invalid_format",
            )


class TestRiskCalculation:
    """Test risk calculation logic."""

    def test_calculate_risks(self, sample_assets, sample_threats):
        """Test risk calculation for assets and threats."""
        results = generator.calculate_risks(sample_assets, sample_threats)

        assert len(results) == 3  # Updated to match sample_threats fixture

        # Results should be sorted by risk level (highest first)
        assert all("risk_level" in result for result in results)
        assert all("likelihood_score" in result for result in results)
        assert all("max_severity" in result for result in results)

        # Check that max severity is correctly calculated
        expected_max_severity = max(asset.severity_score() for asset in sample_assets)
        assert all(
            result["max_severity"] == expected_max_severity for result in results
        )


class TestExporters:
    """Test export functionality."""

    def test_json_export(self, temp_config_dir, tmp_path):
        """Test JSON export functionality."""
        # Load data and calculate risks
        assets, threats, settings, *_ = loader.load_all(temp_config_dir)
        risks = generator.calculate_risks(assets, threats)

        # Export to JSON using simple function instead of class
        from ebiosrm_core import exporters

        output_file = tmp_path / "ebios_risk_assessment.json"
        exporters.export_json(
            {
                "assets": [asset.model_dump() for asset in assets],
                "threats": [threat.model_dump() for threat in threats],
                "risk_results": risks,
                "settings": settings.model_dump(),
            },
            output_file,
        )

        # Verify file exists and contains valid JSON
        assert output_file.exists()

        with open(output_file) as f:
            data = json.load(f)

        assert "metadata" in data
        assert "risks" in data
        assert data["metadata"]["total_risks"] == len(risks)
        assert len(data["risks"]) == len(risks)

    def test_excel_export(self, temp_config_dir, tmp_path):
        """Test Excel export functionality."""
        # Load data and calculate risks
        assets, threats, settings, *_ = loader.load_all(temp_config_dir)
        risks = generator.calculate_risks(assets, threats)

        # Export to Excel using simple function instead of class
        from ebiosrm_core import exporters

        output_file = tmp_path / "ebios_risk_assessment.xlsx"
        exporters.export_excel(
            {
                "assets": [asset.model_dump() for asset in assets],
                "threats": [threat.model_dump() for threat in threats],
                "risk_results": risks,
                "settings": settings.model_dump(),
            },
            output_file,
        )

        # Verify file exists
        assert output_file.exists()
        assert output_file.suffix == ".xlsx"


class TestFullWorkflow:
    """Test complete end-to-end workflow."""

    def test_generator_run_json(self, temp_config_dir, tmp_path):
        """Test complete generator run with JSON output."""
        generator.run(cfg_dir=temp_config_dir, out_dir=tmp_path, fmt="json")

        # Verify output file was created
        output_file = tmp_path / "ebios_risk_assessment.json"
        assert output_file.exists()

        # Verify content structure
        with open(output_file) as f:
            data = json.load(f)

        assert "metadata" in data
        assert "risks" in data
        assert data["metadata"]["total_risks"] > 0

    def test_generator_run_excel(self, temp_config_dir, tmp_path):
        """Test complete generator run with Excel output."""
        generator.run(cfg_dir=temp_config_dir, out_dir=tmp_path, fmt="xlsx")

        # Verify output file was created
        output_file = tmp_path / "ebios_risk_assessment.xlsx"
        assert output_file.exists()

    def test_generator_run_markdown(self, temp_config_dir, tmp_path):
        """Test complete generator run with Markdown output."""
        generator.run(cfg_dir=temp_config_dir, out_dir=tmp_path, fmt="markdown")

        # Verify output file was created
        output_file = tmp_path / "ebios_risk_assessment.md"
        assert output_file.exists()

        # Verify basic markdown structure
        content = output_file.read_text(encoding="utf-8")
        assert "# EBIOS RM Risk Assessment Report" in content
        assert "## Risk Distribution" in content

    def test_unsupported_export_format(self, temp_config_dir, tmp_path):
        """Test error handling for unsupported export format."""
        with pytest.raises(ValueError, match="Unsupported format"):
            generator.run(cfg_dir=temp_config_dir, out_dir=tmp_path, fmt="invalid")


class TestEBIOSRMWorkflows:
    """Test EBIOS RM-specific workflows and data structures."""

    def test_template_integration_with_workflow(self, temp_config_dir, tmp_path):
        """Test l'intégration du nouveau template avec le workflow existant."""
        # Tester uniquement si les modules sont disponibles
        try:
            from scripts.generate_template import EBIOSTemplateGenerator
            from scripts.sync_json_excel import JSONExcelSyncer
        except ImportError:
            pytest.skip("Template generation modules not available")

        # Générer le template
        template_path = tmp_path / "integrated_template.xlsx"
        generator = EBIOSTemplateGenerator()
        generator.generate_template(template_path)

        assert template_path.exists()

        # Synchroniser avec JSON
        json_path = tmp_path / "integrated_schema.json"
        syncer = JSONExcelSyncer(template_path, json_path)
        syncer.sync_excel_to_json()

        assert json_path.exists()

        # Valider cohérence
        issues = syncer.validate_consistency()
        assert len(issues["errors"]) == 0

        # Tester avec le workflow existant
        from ebiosrm_core import generator, loader

        assets, threats, settings, *_ = loader.load_all(temp_config_dir)
        risks = generator.calculate_risks(assets, threats)

        # Vérifier compatibilité des données
        assert len(risks) > 0
        for risk in risks:
            assert "risk_level" in risk
            assert "likelihood_score" in risk
            assert risk["risk_level"] in ["Low", "Medium", "High", "Critical"]

    def test_excel_formulas_integration(self, tmp_path):
        """Test que les formules Excel fonctionnent avec les données réelles."""
        try:
            from scripts.generate_template import EBIOSTemplateGenerator
        except ImportError:
            pytest.skip("Template generation module not available")

        from openpyxl import load_workbook

        # Générer template
        template_path = tmp_path / "formula_test.xlsx"
        generator = EBIOSTemplateGenerator()
        generator.generate_template(template_path)

        # Charger et tester les formules
        wb = load_workbook(template_path)

        # Test Atelier2 - Formules XLOOKUP
        ws = wb["Atelier2_Sources"]

        # Vérifier qu'on a bien des formules dans les bonnes cellules
        for row in range(2, 5):  # Premières lignes
            for col in [2, 3, 4, 5]:  # Colonnes avec formules
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    assert isinstance(
                        cell.value, str
                    ), f"Cellule {cell.coordinate} devrait contenir une formule"
                    assert "XLOOKUP" in cell.value or "IFERROR" in cell.value, (
                        f"Formule manquante dans {cell.coordinate}"
                    )

        wb.close()

    def test_pme_profile_template_compatibility(self, tmp_path):
        """Test la compatibilité du template avec le profil PME."""
        try:
            from scripts.generate_template import EBIOSTemplateGenerator
            from scripts.sync_json_excel import JSONExcelSyncer
        except ImportError:
            pytest.skip("Template generation modules not available")

        # Le template doit supporter les échelles simplifiées PME
        generator = EBIOSTemplateGenerator()
        template_path = tmp_path / "pme_template.xlsx"
        generator.generate_template(template_path)

        # Vérifier que les échelles sont compatibles PME
        json_path = tmp_path / "pme_schema.json"

        syncer = JSONExcelSyncer(template_path, json_path)
        enums = syncer.extract_enums_from_excel()

        # Les échelles doivent avoir 4 niveaux (compatible PME)
        assert len(enums["gravity_scale"]) == 4
        assert len(enums["likelihood_scale"]) == 4

        # Vérifier les libellés français
        gravity_labels = enums["gravity_scale"]
        assert "Négligeable" in gravity_labels
        assert "Critique" in gravity_labels
