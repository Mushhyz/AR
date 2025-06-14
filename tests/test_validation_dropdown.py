"""Tests de validation des listes déroulantes et auto-complétion des mesures."""

import pytest
from pathlib import Path
from openpyxl import load_workbook
from scripts.generate_template import EBIOSTemplateGenerator


class TestDropdownValidation:
    """Tests pour vérifier que les listes déroulantes sont fonctionnelles."""
    
    @pytest.fixture
    def template_path(self, tmp_path):
        """Génère un template de test."""
        template_file = tmp_path / "test_dropdown.xlsx"
        generator = EBIOSTemplateGenerator()
        generator.generate_template(template_file)
        return template_file
    
    def test_dropdown_visible_atelier3(self, template_path):
        """Test que les flèches de listes déroulantes sont visibles dans Atelier 3."""
        wb = load_workbook(template_path)
        ws = wb["Atelier3_Scenarios"]
        
        # Vérifier les validations de données
        validations = ws.data_validations.dataValidation
        assert len(validations) > 0, "Aucune validation de données trouvée"
        
        # Vérifier que showDropDown=False (ce qui force l'affichage)
        dropdown_found = False
        for dv in validations:
            if dv.formula1 and ("=Gravite" in dv.formula1 or "=Vraisemblance" in dv.formula1):
                assert dv.showDropDown == False, f"showDropDown devrait être False pour {dv.formula1}"
                assert dv.showErrorMessage == True, "Message d'erreur non activé"
                assert dv.showInputMessage == True, "Message d'aide non activé"
                dropdown_found = True
        
        assert dropdown_found, "Aucune validation avec flèche trouvée"
    
    def test_dropdown_visible_atelier4(self, template_path):
        """Test que les flèches de listes déroulantes sont visibles dans Atelier 4."""
        wb = load_workbook(template_path)
        ws = wb["Atelier4_Operationnels"]
        
        # Vérifier les validations de données
        validations = ws.data_validations.dataValidation
        assert len(validations) > 0, "Aucune validation de données trouvée"
        
        # Vérifier que showDropDown=False et formules correctes
        measure_validation_found = False
        for dv in validations:
            if dv.formula1 and "=Measure_ID" in dv.formula1:
                assert dv.showDropDown == False, "showDropDown devrait être False pour Measure_ID"
                assert "mesure n'existe pas" in dv.error, "Message d'erreur personnalisé manquant"
                measure_validation_found = True
        
        assert measure_validation_found, "Validation Measure_ID non trouvée"
    
    def test_named_ranges_exist(self, template_path):
        """Test que toutes les plages nommées existent."""
        wb = load_workbook(template_path)
        
        # Plages nommées essentielles pour les validations
        required_ranges = [
            "Gravite", "Vraisemblance", "Valeur_Metier", 
            "Pertinence", "Exposition", "Measure_ID",
            "tbl_Gravite_Valeur", "tbl_Vraisemblance_Valeur", 
            "tbl_Measure_Efficacite", "tbl_Measure_AnnexA"
        ]
        
        for range_name in required_ranges:
            assert range_name in wb.defined_names, f"Plage nommée manquante: {range_name}"
            
            # Vérifier que la plage pointe vers __REFS
            range_obj = wb.defined_names[range_name]
            assert "__REFS" in range_obj.attr_text, f"Plage {range_name} ne pointe pas vers __REFS"


class TestAutoFillMeasures:
    """Tests pour vérifier l'auto-complétion des cellules de mesures."""
    
    @pytest.fixture
    def template_path(self, tmp_path):
        """Génère un template de test."""
        template_file = tmp_path / "test_autofill.xlsx"
        generator = EBIOSTemplateGenerator()
        generator.generate_template(template_file)
        return template_file
    
    def test_autofill_measure_atelier5(self, template_path):
        """Test de l'auto-complétion dans Atelier 5 - Traitement."""
        wb = load_workbook(template_path, data_only=False)
        ws = wb["Atelier5_Traitement"]
        
        # Vérifier les formules d'auto-complétion
        autofill_formulas = []
        for row in range(2, 10):
            # Colonne F - Contrôle AnnexA (XLOOKUP depuis Measure_ID)
            cell_f = ws.cell(row=row, column=6)
            if cell_f.value and isinstance(cell_f.value, str) and cell_f.value.startswith('='):
                assert "XLOOKUP" in cell_f.value, f"Formule XLOOKUP manquante en F{row}"
                assert "Measure_ID" in cell_f.value, f"Référence Measure_ID manquante en F{row}"
                autofill_formulas.append(cell_f.value)
            
            # Colonne J - Efficacité attendue (XLOOKUP depuis catalogue)
            cell_j = ws.cell(row=row, column=10)
            if cell_j.value and isinstance(cell_j.value, str) and cell_j.value.startswith('='):
                assert "XLOOKUP" in cell_j.value, f"Formule XLOOKUP manquante en J{row}"
                assert "tbl_Measure_Efficacite" in cell_j.value, f"Référence efficacité manquante en J{row}"
                autofill_formulas.append(cell_j.value)
            
            # Colonne K - Risque résiduel (calcul avec efficacité)
            cell_k = ws.cell(row=row, column=11)
            if cell_k.value and isinstance(cell_k.value, str) and cell_k.value.startswith('='):
                assert "ISNUMBER" in cell_k.value or "ESTNUM" in cell_k.value, f"Vérification numérique manquante en K{row}"
                autofill_formulas.append(cell_k.value)
        
        assert len(autofill_formulas) > 0, "Aucune formule d'auto-complétion trouvée"
    
    def test_autofill_risk_atelier4(self, template_path):
        """Test de l'auto-complétion du risque dans Atelier 4."""
        wb = load_workbook(template_path, data_only=False)
        ws = wb["Atelier4_Operationnels"]
        
        # Vérifier les formules de calcul de risque
        risk_formulas = []
        for row in range(2, 10):
            # Colonne K - Risque résiduel
            cell_k = ws.cell(row=row, column=11)
            if cell_k.value and isinstance(cell_k.value, str) and cell_k.value.startswith('='):
                assert "IF" in cell_k.value or "SI" in cell_k.value, f"Formule conditionnelle manquante en K{row}"
                risk_formulas.append(cell_k.value)
            
            # Colonne L - Niveau de risque final
            cell_l = ws.cell(row=row, column=12)
            if cell_l.value and isinstance(cell_l.value, str) and cell_l.value.startswith('='):
                assert "Critique" in cell_l.value, f"Calcul niveau critique manquant en L{row}"
                risk_formulas.append(cell_l.value)
        
        assert len(risk_formulas) > 0, "Aucune formule de calcul de risque trouvée"
    
    def test_formula_protection(self, template_path):
        """Test que les formules d'auto-complétion sont protégées et grisées."""
        wb = load_workbook(template_path)
        ws = wb["Atelier5_Traitement"]
        
        # Vérifier les cellules de formules
        protected_cells = 0
        grayed_cells = 0
        
        for row in range(2, 10):
            for col in [6, 9, 10, 11]:  # Colonnes avec formules
                cell = ws.cell(row=row, column=col)
                if (cell.value and isinstance(cell.value, str) and 
                    cell.value.startswith('=') and cell.data_type == "f"):
                    
                    # Vérifier protection
                    if cell.protection.locked:
                        protected_cells += 1
                    
                    # Vérifier grisage
                    if (cell.fill.start_color.rgb and 
                        cell.fill.start_color.rgb.upper() == "D9D9D9".upper()):
                        grayed_cells += 1
        
        assert protected_cells > 0, "Aucune cellule de formule protégée"
        assert grayed_cells > 0, "Aucune cellule de formule grisée"


class TestValidationMessages:
    """Tests pour vérifier les messages d'erreur et d'aide."""
    
    def test_custom_error_messages_atelier3(self, template_path):
        """Test des messages d'erreur personnalisés dans Atelier 3."""
        wb = load_workbook(template_path)
        ws = wb["Atelier3_Scenarios"]
        
        validations = ws.data_validations.dataValidation
        
        custom_messages_found = 0
        for dv in validations:
            if dv.error and dv.prompt:
                if "invalide" in dv.error and "Sélectionnez" in dv.prompt:
                    custom_messages_found += 1
                    assert dv.errorTitle == "Erreur de validation - Atelier 3"
                    assert dv.promptTitle == "Guide de saisie"
        
        assert custom_messages_found >= 3, "Messages personnalisés insuffisants (Gravité, Vraisemblance, Valeur)"
    
    def test_custom_error_messages_atelier4(self, template_path):
        """Test des messages d'erreur personnalisés dans Atelier 4.""" 
        wb = load_workbook(template_path)
        ws = wb["Atelier4_Operationnels"]
        
        validations = ws.data_validations.dataValidation
        
        measure_message_found = False
        for dv in validations:
            if dv.formula1 and "=Measure_ID" in dv.formula1:
                assert "catalogue ISO 27001" in dv.error, "Message spécifique ISO 27001 manquant"
                assert "Annex A" in dv.prompt, "Référence Annex A manquante"
                measure_message_found = True
        
        assert measure_message_found, "Message personnalisé pour mesures non trouvé"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
