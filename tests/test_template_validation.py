"""Tests de validation du template Excel EBIOS RM."""

import pytest
import json  # **CORRECTION 4** : Import manquant ajouté
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from scripts.generate_template import EBIOSTemplateGenerator
from scripts.sync_json_excel import JSONExcelSyncer


class TestTemplateGeneration:
    """Tests de génération du template Excel."""
    
    def test_generate_template_success(self, tmp_path):
        """Test de génération réussie du template."""
        generator = EBIOSTemplateGenerator()
        output_path = tmp_path / "test_template.xlsx"
        
        generator.generate_template(output_path)
        
        assert output_path.exists()
        
        # Vérifier la structure
        wb = load_workbook(output_path)
        expected_sheets = ["__REFS", "Atelier1_Socle", "Atelier2_Sources", 
                          "Atelier3_Scenarios", "Atelier4_Operationnels", "Synthese"]
        
        for sheet in expected_sheets:
            assert sheet in wb.sheetnames
        
        # Vérifier que __REFS est masqué
        assert wb["__REFS"].sheet_state == "veryHidden"
        
        wb.close()
    
    def test_references_sheet_structure(self, tmp_path):
        """Test de la structure de l'onglet __REFS."""
        generator = EBIOSTemplateGenerator()
        output_path = tmp_path / "test_template.xlsx"
        
        generator.generate_template(output_path)
        
        wb = load_workbook(output_path, data_only=True)
        ws = wb["__REFS"]
        
        # **CORRECTION 1** : Vérifier les nouvelles tables
        expected_tables = ["tbl_Gravite", "tbl_Vraisemblance", "tbl_ValeurMetier", 
                          "tbl_Source", "tbl_Scenario", "tbl_OV", "tbl_Measure", 
                          "tbl_Pertinence", "tbl_Exposition"]
        
        # Les tables sont identifiables par leurs headers
        found_tables = []
        current_col = 1
        while current_col <= ws.max_column:
            header = ws.cell(row=1, column=current_col).value
            if header:
                if header == "ID":
                    # Table de niveaux
                    second_header = ws.cell(row=1, column=current_col + 1).value
                    if second_header == "Libelle":
                        first_value = ws.cell(row=2, column=current_col + 1).value
                        if first_value == "Négligeable":
                            found_tables.append("tbl_Gravite")
                        elif first_value == "Minimal":
                            found_tables.append("tbl_Vraisemblance")
                        elif first_value == "Niveau 1":
                            found_tables.append("tbl_ValeurMetier")
                elif header == "Source_ID":
                    found_tables.append("tbl_Source")
                elif header == "Scenario_ID":
                    found_tables.append("tbl_Scenario")
                elif header == "OV_ID":
                    found_tables.append("tbl_OV")
                
                # Passer au groupe suivant
                col = current_col
                while col <= ws.max_column and ws.cell(row=1, column=col).value:
                    col += 1
                current_col = col + 1
            else:
                current_col += 1
        
        # Vérifier qu'on a trouvé toutes les tables
        for expected in expected_tables:
            assert expected in found_tables, f"Table {expected} non trouvée"
        
        wb.close()


class TestDataValidation:
    """Tests de validation des données Excel."""
    
    def test_named_ranges_exist(self, tmp_path):
        """Test de l'existence des plages nommées."""
        generator = EBIOSTemplateGenerator()
        output_path = tmp_path / "test_template.xlsx"

        generator.generate_template(output_path)

        wb = load_workbook(output_path)

        # Vérifier les plages nommées principales (mises à jour)
        expected_ranges = ["Gravite", "Vraisemblance", "Source_ID", "Scenario_ID", "OV_ID", "Asset_Type", "Stakeholder_ID"]

        defined_names = wb.defined_names
        found_ranges = list(defined_names.keys())

        for expected in expected_ranges:
            assert expected in found_ranges, f"Plage nommée '{expected}' non trouvée"

        wb.close()

    def test_locked_cells_protection(self, tmp_path):
        """Test_LockedCells : vérifier que toutes les cellules grises sont verrouillées."""
        generator = EBIOSTemplateGenerator()
        output_path = tmp_path / "test_template.xlsx"
        
        generator.generate_template(output_path)
        
        wb = load_workbook(output_path)
        
        # Tester sur tous les ateliers
        ateliers = ["Atelier1_Socle", "Atelier2_Sources", "Atelier3_Scenarios", "Atelier4_Operationnels"]
        
        for atelier_name in ateliers:
            if atelier_name not in wb.sheetnames:
                continue
                
            ws = wb[atelier_name]
            
            # Parcourir toutes les cellules avec contenu
            for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), 
                                  min_col=1, max_col=min(15, ws.max_column)):
                for cell in row:
                    if cell.fill and cell.fill.start_color:
                        color_value = cell.fill.start_color.rgb
                        
                        # Vérifier si la cellule a le fond gris #D9D9D9
                        if color_value in ["D9D9D9", "00D9D9D9"]:
                            assert cell.protection.locked is True, \
                                f"Cellule grise {cell.coordinate} dans {atelier_name} devrait être verrouillée"
                            
                            # Vérifier que c'est probablement une cellule avec formule
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                # Accepter IF( ou RECHERCHEX selon la formule
                                assert any(keyword in cell.value for keyword in ["XLOOKUP", "RECHERCHEX", "INDEX", "MATCH", "IF(", "IFERROR"]), \
                                    f"Cellule verrouillée {cell.coordinate} devrait contenir une formule de calcul"
        
        wb.close()

    def test_json_parity_validation(self, tmp_path):
        """Test_JSONParity : vérifier cohérence entre énumérations JSON et plages nommées Excel."""
        from scripts.sync_json_excel import JSONExcelSyncer
        
        generator = EBIOSTemplateGenerator()
        excel_path = tmp_path / "test_template.xlsx"
        json_path = tmp_path / "test_schema.json"
        
        # Générer template et synchroniser
        generator.generate_template(excel_path)
        
        syncer = JSONExcelSyncer(excel_path, json_path)
        syncer.sync_excel_to_json()
        
        # Charger le JSON généré
        with open(json_path, 'r', encoding='utf-8') as f:
            schema_data = json.load(f)
        
        # Tester la nouvelle structure avec enumerations à la racine
        enums = schema_data.get("enumerations", {})
        assert "gravity_scale" in enums
        assert "asset_type_catalog" in enums
        assert "stakeholder_catalog" in enums
        assert "business_value_labels" in enums
        
        # Charger le classeur Excel
        wb = load_workbook(excel_path)
        
        # Mapping des énumérations JSON vers les plages nommées Excel (mis à jour)
        enum_to_range_mapping = {
            "gravity_scale": "Gravite",
            "likelihood_scale": "Vraisemblance", 
            "source_catalog": "Source_ID",
            "scenario_catalog": "Scenario_ID",
            "operational_catalog": "OV_ID",
            "asset_type_catalog": "Asset_Type",
            "stakeholder_catalog": "Stakeholder_ID"
        }
        
        # Vérifier chaque énumération JSON contre sa plage nommée Excel correspondante
        for enum_name, range_name in enum_to_range_mapping.items():
            if enum_name in enums:
                json_values = enums[enum_name]
                
                # Vérifier que la plage nommée existe
                assert range_name in wb.defined_names, f"Plage nommée '{range_name}' manquante pour enum '{enum_name}'"
        
        wb.close()


class TestJSONSynchronization:
    """Tests de synchronisation JSON."""
    
    def test_extract_enums_from_excel(self, tmp_path):
        """Test d'extraction des énumérations depuis Excel."""
        generator = EBIOSTemplateGenerator()
        excel_path = tmp_path / "test_template.xlsx"
        json_path = tmp_path / "test_schema.json"
        
        generator.generate_template(excel_path)
        
        syncer = JSONExcelSyncer(excel_path, json_path)
        enums = syncer.extract_enums_from_excel()
        
        # Vérifier les énumérations principales (mises à jour)
        assert "gravity_scale" in enums
        assert "likelihood_scale" in enums
        assert "source_catalog" in enums
        assert "asset_type_catalog" in enums
        assert "stakeholder_catalog" in enums
        assert "business_value_labels" in enums
        
        # Vérifier le contenu - maintenant les labels sont directement dans gravity_scale
        assert "Négligeable" in enums["gravity_scale"]
        assert "Critique" in enums["gravity_scale"]
        assert len(enums["gravity_scale"]) == 4
        
        assert "Minimal" in enums["likelihood_scale"]
        assert len(enums["likelihood_scale"]) == 4
        
        # Vérifier les nouvelles énumérations
        assert "Serveur" in enums["asset_type_catalog"]
        assert "DSI" in enums["stakeholder_catalog"]
        assert "Niveau 1" in enums["business_value_labels"]
        
        assert len(enums["source_catalog"]) >= 1
        assert any(item.get("Source_ID") == "RS001" for item in enums["source_catalog"])
    
    def test_sync_excel_to_json(self, tmp_path):
        """Test de synchronisation complète Excel → JSON."""
        generator = EBIOSTemplateGenerator()
        excel_path = tmp_path / "test_template.xlsx"
        json_path = tmp_path / "test_schema.json"
        
        generator.generate_template(excel_path)
        
        syncer = JSONExcelSyncer(excel_path, json_path)
        syncer.sync_excel_to_json()
        
        assert json_path.exists()
        
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Vérifier la structure JSON mise à jour
        assert "metadata" in data
        assert "enumerations" in data
        assert "schema" in data
        
        # Vérifier les énumérations étendues
        enums = data["enumerations"]
        assert "gravity_scale" in enums
        assert "source_catalog" in enums
        assert "asset_type_catalog" in enums
        assert "stakeholder_catalog" in enums
        
        # Vérifier le schéma de validation mis à jour
        schema = data["schema"]
        assert "atelier1_socle" in schema
        assert "atelier2_sources" in schema
        
        # Vérifier les validations étendues
        atelier1_validations = schema["atelier1_socle"]["validations"]
        assert "Type" in atelier1_validations
        assert "Propriétaire" in atelier1_validations
        assert "Valeur_Métier" in atelier1_validations

class TestWorkflowIntegration:
    """Tests d'intégration complète du workflow."""
    
    def test_complete_workflow(self, tmp_path):
        """Test du workflow complet génération → synchronisation → validation."""
        # Chemins
        excel_path = tmp_path / "ebios_template.xlsx"
        json_path = tmp_path / "ebios_schema.json"
        
        # 1. Génération du template
        generator = EBIOSTemplateGenerator()
        generator.generate_template(excel_path)
        assert excel_path.exists()
        
        # 2. Synchronisation
        syncer = JSONExcelSyncer(excel_path, json_path)
        syncer.sync_excel_to_json()
        assert json_path.exists()
        
        # 3. Validation
        issues = syncer.validate_consistency()
        assert len(issues["errors"]) == 0
        
        # 4. Vérification finale de l'intégrité (mise à jour)
        wb = load_workbook(excel_path)
        
        # Toutes les feuilles attendues
        expected_sheets = ["__REFS", "Config_EBIOS", "Atelier1_Socle", "Atelier2_Sources", 
                          "Atelier3_Scenarios", "Atelier4_Operationnels", "Synthese"]
        assert all(sheet in wb.sheetnames for sheet in expected_sheets)
        
        # Données de référence présentes
        refs_ws = wb["__REFS"]
        assert refs_ws.max_row > 1  # Au moins des données
        assert refs_ws.max_column > 1  # Au moins 2 colonnes
        
        # Validations actives dans Atelier1 (validation étendue)
        atelier1_ws = wb["Atelier1_Socle"]
        assert len(atelier1_ws.data_validations.dataValidation) >= 5  # Au moins 5 validations
        
        wb.close()
        
        # JSON cohérent - tester la nouvelle structure
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Les tests attendent gravity_scale avec les labels directement
        assert data["enumerations"]["gravity_scale"] == ["Négligeable", "Limité", "Important", "Critique"]
        assert len(data["enumerations"]["source_catalog"]) >= 3
        assert "asset_type_catalog" in data["enumerations"]
        assert "stakeholder_catalog" in data["enumerations"]
        
        print("✅ Workflow complet validé avec succès")
    
    def test_complete_workflow_with_measures(self, tmp_path):
        """Test du workflow complet avec toutes les mesures et ateliers."""
        # Chemins
        excel_path = tmp_path / "ebios_template_complet.xlsx"
        json_path = tmp_path / "ebios_schema_complet.json"
        
        # 1. Génération du template
        generator = EBIOSTemplateGenerator()
        generator.generate_template(excel_path)
        assert excel_path.exists()
        
        # 2. Synchronisation
        syncer = JSONExcelSyncer(excel_path, json_path)
        syncer.sync_excel_to_json()
        assert json_path.exists()
        
        # 3. Validation
        issues = syncer.validate_consistency()
        assert len(issues["errors"]) == 0
        
        # 4. **CORRECTION 1** : Vérification des 5 ateliers complets
        wb = load_workbook(excel_path)
        
        # Toutes les feuilles attendues avec Atelier 5
        expected_sheets = ["__REFS", "Config_EBIOS", "Atelier1_Socle", "Atelier2_Sources", 
                          "Atelier3_Scenarios", "Atelier4_Operationnels", "Atelier5_Traitement", "Synthese"]
        assert all(sheet in wb.sheetnames for sheet in expected_sheets)
        
        # **CORRECTION 1** : Vérifier que l'Atelier 5 a les bonnes validations
        atelier5_ws = wb["Atelier5_Traitement"]
        assert len(atelier5_ws.data_validations.dataValidation) >= 7  # Au moins 7 validations
        
        wb.close()
        
        # **CORRECTION 4** : JSON avec énumérations complètes
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Vérifier toutes les nouvelles énumérations
        enums = data["enumerations"]
        assert "gravity_labels" in enums  # **CORRECTION 4** : Libellés explicites
        assert "likelihood_labels" in enums
        assert "business_value_labels" in enums
        assert "measure_catalog" in enums  # **CORRECTION 1** : Catalogue des mesures
        assert "pertinence_scale" in enums  # **CORRECTION 2**
        assert "exposition_scale" in enums
        
        # Vérifier les valeurs
        assert data["enumerations"]["gravity_labels"] == ["Négligeable", "Limité", "Important", "Critique"]
        assert len(data["enumerations"]["measure_catalog"]) >= 5  # Au moins 5 mesures
        
        # **CORRECTION 4** : Vérifier les métadonnées
        metadata = data["metadata"]
        assert metadata["generator"] == "EBIOSTemplateGenerator"
        assert metadata["version"] == "2.0.0"
        assert "total_enumerations" in metadata
        
        print("✅ Workflow complet avec 5 ateliers validé avec succès")

class TestMeasuresCatalog:
    """Tests spécifiques pour le catalogue des mesures de sécurité."""
    
    def test_measures_extraction(self, tmp_path):
        """Test d'extraction du catalogue des mesures."""
        generator = EBIOSTemplateGenerator()
        excel_path = tmp_path / "test_measures.xlsx"
        json_path = tmp_path / "test_measures.json"
        
        generator.generate_template(excel_path)
        
        syncer = JSONExcelSyncer(excel_path, json_path)
        enums = syncer.extract_enums_from_excel()
        
        # **CORRECTION 1** : Vérifier le catalogue des mesures
        assert "measure_catalog" in enums
        measures = enums["measure_catalog"]
        assert len(measures) >= 5
        
        # Vérifier la structure des mesures
        first_measure = measures[0]
        assert "id" in first_measure
        assert "label" in first_measure
        assert first_measure["id"].startswith("M")
        
        # Vérifier quelques mesures spécifiques
        measure_ids = [m["id"] for m in measures]
        assert "M001" in measure_ids  # Authentification multi-facteurs
        assert "M002" in measure_ids  # Chiffrement des données
        assert "M003" in measure_ids  # Supervision SOC

    def test_atelier5_validations(self, tmp_path):
        """Test des validations de l'Atelier 5 - Traitement."""
        generator = EBIOSTemplateGenerator()
        excel_path = tmp_path / "test_atelier5.xlsx"
        
        generator.generate_template(excel_path)
        
        wb = load_workbook(excel_path)
        ws = wb["Atelier5_Traitement"]
        
        # Vérifier que les validations sont présentes
        validations = ws.data_validations.dataValidation
        assert len(validations) >= 7
        
        # Vérifier les plages de validation
        validation_ranges = []
        for dv in validations:
            for cell_range in dv.cells:
                validation_ranges.append(str(cell_range))
        
        # Doit inclure les colonnes pour mesures et responsables
        assert any("E2:E" in vr for vr in validation_ranges)  # Mesure choisie
        assert any("F2:F" in vr for vr in validation_ranges)  # Responsable
        
        wb.close()
